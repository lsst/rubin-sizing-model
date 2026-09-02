#!/usr/bin/env python3
"""
Rubin Observatory USDF Sizing Model Generator

Reads tunable parameters from sizing_params.yaml and generates an Excel
spreadsheet with live formulas, charts, and a DR readiness summary.
The generated xlsx uses Excel formulas so that manual edits inside the
spreadsheet propagate automatically.

Usage:
    python generate_sizing_model.py [sizing_params.yaml] [output.xlsx]
"""

import math
import os
import sys
import yaml
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.chart import AreaChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

# ── Helpers ──────────────────────────────────────────────────────────────────

def col(loy_index):
    """Return column letter for LOY index (0-based: LOY1=C, LOY2=D, …)."""
    return get_column_letter(loy_index + 3)  # A=1, B=2, C=3


def col_abs(loy_index):
    """Absolute column reference like $C."""
    return f"${col(loy_index)}"


def cell(sheet_prefix, row, loy_index):
    """Return a cell reference like 'Ops Storage'!C5."""
    c = col(loy_index)
    if sheet_prefix:
        return f"'{sheet_prefix}'!{c}{row}"
    return f"{c}{row}"


def acell(row, loy_index):
    """Local cell reference like C5."""
    return f"{col(loy_index)}{row}"


def hdr_fill():
    return PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")


def hdr_font():
    return Font(bold=True, color="FFFFFF", size=11)


def section_font():
    return Font(bold=True, size=11, color="1F4E79")


def num_fmt_tb():
    return "#,##0.0"


def num_fmt_int():
    return "#,##0"


def num_fmt_dollars():
    return "$#,##0"


def num_fmt_millions():
    return "$#,##0.000"


def apply_header_style(ws, row, max_col):
    for c in range(1, max_col + 1):
        cl = ws.cell(row=row, column=c)
        cl.fill = hdr_fill()
        cl.font = hdr_font()
        cl.alignment = Alignment(horizontal="center", wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fleet_batch(P):
    """Authoritative Slurm batch hardware specs (no duplicate costs.* fields)."""
    return P["current_fleet"]["batch"]


def eff_torino_cores(P):
    """Effective cores per new batch node (Torino class)."""
    return int(P["current_fleet"]["batch"]["torino_cores_per_node"])


# ── MAIN ─────────────────────────────────────────────────────────────────────

def generate_png_charts(P, N, FY0):
    """Generate standalone PNG chart files via matplotlib."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.ticker as mticker
    except ImportError:
        print("matplotlib not installed — skipping PNG chart generation.")
        return

    os.makedirs("charts", exist_ok=True)

    obs = P["observing"]
    img = P["imaging"]
    drp = P["drp"]
    cat = P["catalogs"]
    qd = P["qserv_data_per_node"]
    retention = P.get("storage_retention", {})
    qserv_window = retention.get("qserv_window_years", 3)

    loy_labels = [f"LOY{i+1}" for i in range(N)]
    years = [FY0 + i for i in range(N)]

    input_tb_year = (obs["visits_per_night"] * obs["nights_per_year"]
                     * img["lsstcam_image_size_tb"] * img["raw_compression"])
    ch_per_tb = drp["core_hours_per_input_tb"]
    carryover = drp.get("commissioning_carryover_ch", 0)
    proc_hrs = drp["processing_window_days"] * 24
    sci_imgs_year = (obs["nights_per_year"] * obs["visits_per_night"]
                     * obs["images_per_visit"])

    # ── Compute data arrays ──────────────────────────────────────────
    drp_ch = []
    raw_tb, out_tb, parq_tb, coadd_tb = [], [], [], []
    qserv_czar, qserv_db = [], []
    fast_pb, normal_pb, qserv_pb, objstore_pb, tape_pb = [], [], [], [], []

    for i in range(N):
        yr = i + 1
        ch = input_tb_year * yr * ch_per_tb + (carryover if i == 0 else 0)
        drp_ch.append(ch)

        o_imgs = img["output_image_tb_per_visit"] * sci_imgs_year / obs["images_per_visit"]
        pq = img["parquet_tb_per_visit"] * sci_imgs_year / obs["images_per_visit"]

        objs = cat["objects_per_year"][i]
        srcs = cat["sources_per_year"][i]
        forced = cat["forced_sources_base_per_year"][i]
        obj_row = cat["object_row_bytes"]
        obj_extra = cat["object_extra_row_bytes"]
        src_row = cat["source_row_bytes"]
        f_row = cat["forced_source_row_bytes"]
        rep = P["qserv"]["replication_factor"]

        qc = objs * obj_row * rep / 1e12
        qdb = (objs * (obj_row + obj_extra) + srcs * src_row + forced * f_row) * rep / 1e12

        eng_imgs = obs["engineering_fraction"] * sci_imgs_year
        cal_imgs = obs["nights_per_year"] * obs["calibration_images_per_day"]
        cal_comp = img["calibration_compression"]
        ri = (img["lsstcam_image_size_tb"] * img["raw_compression"]
              * (sci_imgs_year + eng_imgs)
              + img["lsstcam_image_size_tb"] * cal_comp * cal_imgs)

        prec = P["precursor"]
        coadd_per_dr = (prec["hsc_rc2_output_coadd_tb"] / prec["hsc_rc2_area_deg2"]
                        * prec["lsstcam_area_deg2"] * prec["coadd_products"])

        raw_tb.append(ri)
        out_tb.append(o_imgs)
        parq_tb.append(pq)
        coadd_tb.append(round(coadd_per_dr, 1))
        qserv_czar.append(qc)
        qserv_db.append(qdb)

    # Accumulate on-floor storage tiers
    prompt_t1 = P["prompt"]["tier1_tb_per_night"] * obs["nights_per_year"]
    prompt_t2 = P["prompt"]["tier2_tb_per_night"] * P["prompt"]["tier2_retention_days"]
    users_count = P["users"]["count_per_year"]
    users_tb = P["users"]["tb_per_user_per_year"]

    for i in range(N):
        sf = P["storage_fractions"]
        apdb = sf["apdb_reference_tb"] / sf["apdb_reference_visits"] * obs["visits_per_night"] * obs["nights_per_year"]
        # Qserv Czar 3yr window
        qc_window = sum(qserv_czar[max(0, i - 2):i + 1])
        fast_i = (apdb + qc_window + prompt_t2) / 1000
        fast_pb.append(fast_i)

        scratch = sf["scratch_fraction"] * out_tb[i]
        user_h = users_count[i] * users_tb[i]
        sub = (out_tb[i] + parq_tb[i] + sf["sims_output_tb"] + scratch + qserv_czar[i]
               + qserv_db[i] + user_h + apdb + prompt_t1 + prompt_t2
               + coadd_tb[i] + raw_tb[i])
        other = sf["misc_overhead_fraction"] * sub
        normal_i = (out_tb[i] + parq_tb[i] + sf["sims_output_tb"] + scratch
                    + qserv_czar[i] + qserv_db[i] + user_h + other) / 1000
        normal_pb.append(normal_i)

        # Qserv Storage 3yr window
        qs_window = sum(qserv_db[max(0, i - 2):i + 1])
        qserv_pb.append(qs_window / 1000)

        # Object Store
        raw_cum = sum(raw_tb[:i + 1])
        lossy = img["lossy_compression"]
        oi_start = max(0, i - 1)
        out_window = sum(out_tb[oi_start:i + 1])
        coadd_start = max(0, i - 2)
        coadd_window = sum(coadd_tb[coadd_start:i + 1])
        parq_start = max(0, i - 2)
        parq_window = sum(parq_tb[parq_start:i + 1])
        t1_cum = prompt_t1 * (i + 1)
        os_i = (raw_cum + out_window * lossy + coadd_window
                + parq_window + t1_cum) / 1000
        objstore_pb.append(os_i)

        # Tape
        data_backup = normal_i * 1000 - scratch - qserv_czar[i] - qserv_db[i]
        tape_i = (raw_cum + sum(coadd_tb[:i + 1]) + data_backup) / 1000
        tape_pb.append(tape_i)

    # ── Chart 1: Storage by Tier ─────────────────────────────────────
    fig, ax = plt.subplots(figsize=(12, 7))
    ax.plot(loy_labels, fast_pb, "o-", label="Fast (NVMe)")
    ax.plot(loy_labels, normal_pb, "s-", label="Normal (HDD)")
    ax.plot(loy_labels, qserv_pb, "^-", label="Qserv")
    ax.plot(loy_labels, objstore_pb, "D-", label="Object Store")
    ax.plot(loy_labels, tape_pb, "v-", label="Tape")
    ax.set_ylabel("Petabytes")
    ax.set_xlabel("Year of Operations")
    ax.set_title("USDF Storage Forecast by Tier")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig("charts/storage_by_tier.png", dpi=150)
    plt.close(fig)

    # ── Chart 2: Detail tiers (Fast, Qserv, Object Store) ───────────
    fig, ax = plt.subplots(figsize=(12, 7))
    ax.plot(loy_labels, fast_pb, "o-", label="Flash (NVMe)")
    ax.plot(loy_labels, qserv_pb, "^-", label="Qserv")
    ax.plot(loy_labels, objstore_pb, "D-", label="Object Store")
    ax.set_ylabel("Petabytes")
    ax.set_xlabel("Year of Operations")
    ax.set_title("Storage Detail: Flash / Qserv / Object Store")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig("charts/storage_detail.png", dpi=150)
    plt.close(fig)

    # ── Chart 3: Data Product Breakdown (stacked area) ───────────────
    raw_cum_pb = [sum(raw_tb[:i + 1]) / 1000 for i in range(N)]
    out_cum_pb = [out_tb[i] / 1000 for i in range(N)]
    coadd_pb = [c / 1000 for c in coadd_tb]
    parq_pb = [p / 1000 for p in parq_tb]
    catdb_pb = [(qserv_czar[i] + qserv_db[i]) / 1000 for i in range(N)]

    fig, ax = plt.subplots(figsize=(12, 7))
    ax.stackplot(loy_labels, raw_cum_pb, out_cum_pb, coadd_pb, parq_pb, catdb_pb,
                 labels=["Raw Images", "Processed Images", "Co-added Images",
                         "Parquet Tables", "Catalog Database"],
                 alpha=0.8)
    ax.set_ylabel("Petabytes")
    ax.set_xlabel("Year of Operations")
    ax.set_title("Data Product Size Breakdown")
    ax.legend(loc="upper left")
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig("charts/data_product_breakdown.png", dpi=150)
    plt.close(fig)

    # ── Chart 4: DRP Core-Hours ──────────────────────────────────────
    fig, ax = plt.subplots(figsize=(12, 7))
    ax.plot(loy_labels, [c / 1e6 for c in drp_ch], "o-", color="#2563eb",
            linewidth=2, markersize=6, label="DRP Core-Hours")
    ax.set_ylabel("Millions of Core-Hours")
    ax.set_xlabel("Year of Operations")
    ax.set_title("DRP Core-Hours Projection")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:,.0f}M"))
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig("charts/drp_core_hours.png", dpi=150)
    plt.close(fig)

    # ── Chart 5: Total Online Storage vs Installed ───────────────────
    online_pb = [fast_pb[i] + normal_pb[i] + objstore_pb[i] + qserv_pb[i]
                 for i in range(N)]
    installed = P["current_fleet"]["storage"]["total_installed_pb"]

    fig, ax = plt.subplots(figsize=(12, 7))
    ax.plot(loy_labels, online_pb, "s-", color="#dc2626", linewidth=2, label="Online Storage Needed")
    ax.axhline(y=installed, color="#16a34a", linestyle="--", linewidth=2, label=f"Installed ({installed} PB)")
    ax.fill_between(loy_labels, installed, online_pb, alpha=0.15, color="#dc2626")
    ax.set_ylabel("Petabytes")
    ax.set_xlabel("Year of Operations")
    ax.set_title("Online Storage: Needed vs Installed")
    ax.legend()
    ax.grid(True, alpha=0.3)
    fig.tight_layout()
    fig.savefig("charts/storage_gap.png", dpi=150)
    plt.close(fig)

    print(f"Saved charts/ (5 PNG files)")


def main():
    yaml_path = sys.argv[1] if len(sys.argv) > 1 else "sizing_params.yaml"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "rubin_sizing_model_2026.xlsx"

    with open(yaml_path) as f:
        P = yaml.safe_load(f)

    N = P["general"]["num_loy"]
    FY0 = P["general"]["start_fiscal_year"]

    wb = openpyxl.Workbook()

    # Create sheets in order
    ws_ref = wb.active
    ws_ref.title = "Reference Data"
    ws_os = wb.create_sheet("Ops Storage")
    ws_oc = wb.create_sheet("Ops Compute")
    ws_model = wb.create_sheet("Model")
    ws_costs = wb.create_sheet("Ops Costs")
    ws_pp = wb.create_sheet("Purchase Plan")
    ws_intl = wb.create_sheet("International Compute")
    ws_yr = wb.create_sheet("Yearly Readiness")
    ws_hr = wb.create_sheet("Hardware Refresh")
    ws_charts = wb.create_sheet("Charts")

    last_col = N + 2  # column index of LOY10

    build_reference_tab(ws_ref, P, N, FY0)
    build_model_tab(ws_model, P, N, FY0)
    build_ops_storage_tab(ws_os, P, N, FY0)
    build_ops_compute_tab(ws_oc, P, N, FY0)
    build_ops_costs_tab(ws_costs, P, N, FY0)
    build_purchase_plan_tab(ws_pp, P, N, FY0)
    build_international_tab(ws_intl, P, N, FY0)
    build_yearly_readiness_tab(ws_yr, P, N, FY0)
    build_hardware_refresh_tab(ws_hr, P, N, FY0)
    build_charts_tab(ws_charts, P, N, FY0)

    # ── Define Named Ranges ──────────────────────────────────────────
    define_named_ranges(wb, P)

    wb.save(output_path)
    print(f"Saved {output_path}")

    # ── Generate PNG charts via matplotlib ─────────────────────────
    generate_png_charts(P, N, FY0)

    # ── Lean communication workbook (same params, never diverges) ──
    from generate_comm_workbook import build_comm_workbook
    build_comm_workbook(P, N, FY0, "rubin_usdf_model_2026_condensed.xlsx")

    # ── Pricing edition (2026 capture unit costs) ──────────────────
    from generate_pricing_workbook import build_pricing_workbook
    build_pricing_workbook(P, N, FY0, "rubin_usdf_model_2026_condensed_pricing.xlsx")


# ══════════════════════════════════════════════════════════════════════
# Reference Data Tab
# ══════════════════════════════════════════════════════════════════════

def build_reference_tab(ws, P, N, FY0):
    ws.sheet_properties.tabColor = "70AD47"
    set_col_widths(ws, [35, 20, 50])

    row = 1
    ws.cell(row=row, column=1, value="Rubin USDF Sizing Model — Reference Data").font = Font(bold=True, size=14)
    row += 2
    ws.cell(row=row, column=1, value="Parameter").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Value").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Source / Notes").font = Font(bold=True)
    apply_header_style(ws, row, 3)

    def add(label, value, note=""):
        nonlocal row
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=note)

    row += 1
    ws.cell(row=row, column=1, value="DRP Estimates").font = section_font()
    add("DR1 node-days (reference only)", P["drp"]["dr1_node_days"],
        "Not used in formulas; see core_hours_per_input_tb")
    add("DRP CPU doc reference", P["drp"].get("document_cpu_reference", ""), "DRP Resource Usage Estimates")
    add("Safety margin", P["drp"]["safety_margin"])
    add("Cores per node", P["drp"]["cores_per_node"])
    add("Processing window (days)", P["drp"]["processing_window_days"])
    add("DR1 total storage (PB)", P["drp"]["dr1_storage_total_pb"], "Compressed")
    add("DR1 release products (PB)", P["drp"]["dr1_storage_release_pb"])
    add("Wall-to-CPU ratio", P["drp"]["wall_to_cpu_ratio"], "Task-level; see idle multiplier")
    add(
        "DRP idle / inter-stage core multiplier",
        P["drp"].get("idle_and_inter_stage_multiplier", 1.0),
        "Calendar / gap overhead on concurrent DRP cores (K-T; CM to refine)",
    )
    dp = P.get("developer_pilot") or {}
    add(
        "Developer-pilot node-days / year",
        dp.get("additional_node_days_per_year", 0),
        "Added to DRP core-hours (K-T; confirm with Yusra/CM)",
    )
    add(
        "USDF-RSP extra fraction of DRP concurrent",
        dp.get("usdf_rsp_extra_fraction_of_drp_concurrent", 0),
        "Applied after idle multiplier on DRP cores row",
    )
    add("DR2 multiplier", P["drp"]["dr2_multiplier"])

    row += 1
    ws.cell(row=row, column=1, value="Prompt Products (RFC-1134)").font = section_font()
    add("Tier 1 TB/night", P["prompt"]["tier1_tb_per_night"], "Long-term retention until DR1")
    add("Tier 2 TB/night", P["prompt"]["tier2_tb_per_night"], "30-day rolling")
    add("DR1 target year", P["prompt"]["dr1_target_year"])

    row += 1
    ws.cell(row=row, column=1, value="Qserv (from Igor)").font = section_font()
    add("Current nodes", P["qserv"]["current_nodes"])
    add("Cores per node", P["qserv"]["cores_per_node"])
    add("Storage TB per node", P["qserv"]["storage_tb_per_node"])
    add("Replication factor", P["qserv"]["replication_factor"])

    row += 1
    ws.cell(row=row, column=1, value="Current Fleet (2026 — Rubin USDF)").font = section_font()
    fleet = P["current_fleet"]
    add("Milano batch nodes", fleet["batch"]["milano_nodes"], "FY24 CPU Planner, Rubin USDF")
    add("Torino batch nodes", fleet["batch"]["torino_nodes"], "Confirmed available to Rubin")
    add("K8s nodes", fleet["k8s"]["nodes"], "Initial k8s Capture, Rubin USDF")
    add("Interactive nodes", fleet["interactive"]["nodes"], "Initial InteractiveWorkflow Cap")
    add("GPU H200 nodes", fleet["gpu"]["h200_nodes"], "Initial GPU Capture, Rubin USDF")
    add("Qserv nodes", fleet["qserv"]["nodes"], "From conversation with Igor")
    add("Cassandra nodes", fleet.get("cassandra", {}).get("nodes", ""), "Rubin Hardware Summary Mar 2025 JTM (Yemi)")
    add("NVMe servers", fleet.get("nvme_servers", {}).get("nodes", ""), "Weka / flash tier (JTM)")
    add("JBOD enclosures (total / deployed)", (
        f'{fleet.get("jbods", {}).get("total_enclosures", "")} / '
        f'{fleet.get("jbods", {}).get("deployed_enclosures", "")}'
    ), "JTM")
    add("Installed storage (PB)", fleet["storage"]["total_installed_pb"], "Initial Storage Capture, Rubin USDF")

    total_batch_cores = (
        fleet["batch"]["milano_nodes"] * fleet["batch"]["milano_cores_per_node"] +
        fleet["batch"]["torino_nodes"] * fleet["batch"]["torino_cores_per_node"]
    )
    add("Total batch cores", total_batch_cores, "Milano + Torino")

    row += 1
    ws.cell(row=row, column=1, value="Node Costs").font = section_font()
    costs = P["costs"]["nodes"]
    add("Milano node (USD)", costs["milano_usd"])
    add("Torino node (USD)", costs["torino_usd"], "Estimated")
    add("K8s node (USD)", costs["k8s_usd"])
    add("Qserv node (USD)", costs["qserv_usd"])
    add("GPU H200 node (USD)", costs["gpu_h200_usd"])

    row += 1
    ws.cell(row=row, column=1, value="Storage Costs ($/TB)").font = section_font()
    sc = P["costs"]["storage"]
    add("Flash (NVMe/Weka)", sc["flash_per_tb"])
    add("Normal (SATA HDD/Ceph)", sc["normal_per_tb"])
    add("Latent (Object Store)", sc["latent_per_tb"])
    add("Tape (LTO)", sc["tape_per_tb"])

    row += 1
    ws.cell(row=row, column=1, value="Lifecycle").font = section_font()
    add("Compute lifetime (years)", P["lifecycle"]["compute_lifetime_years"])
    add("Storage lifetime (years)", P["lifecycle"]["storage_lifetime_years"])

    row += 1
    ws.cell(row=row, column=1, value="Price Decrease Factors (annual)").font = section_font()
    pf = P["price_factors"]
    add("CPU", pf["cpu_annual_decrease"])
    add("Storage", pf["storage_annual_decrease"])
    add("Qserv", pf["qserv_annual_decrease"])


# ══════════════════════════════════════════════════════════════════════
# Model Tab
# ══════════════════════════════════════════════════════════════════════

def build_model_tab(ws, P, N, FY0):
    ws.sheet_properties.tabColor = "FFC000"
    set_col_widths(ws, [38, 18, 18, 18, 18, 14, 14])

    # ── Machine Catalog (rows 1-10) ──────────────────────────────────
    r = 1
    ws.cell(row=r, column=1, value="Machine Catalog").font = Font(bold=True, size=13)
    r = 2
    for ci, hdr in enumerate(["Type", "Cores", "RAM (GB)", "Eff Cores/Node", "Cost (USD)", "Purpose"], 1):
        ws.cell(row=r, column=ci, value=hdr)
    apply_header_style(ws, r, 6)

    cn = P["costs"]["nodes"]
    fb = fleet_batch(P)
    machines = [
        ("Milano (batch)", fb["milano_cores_per_node"], fb["milano_ram_gb"],
         fb["milano_cores_per_node"], cn["milano_usd"], "Current Slurm batch"),
        ("Torino (batch)", fb["torino_cores_per_node"], fb["torino_ram_gb"],
         fb["torino_cores_per_node"], cn["torino_usd"], "Future Slurm batch"),
        ("K8s node", cn["k8s_cores"], P["current_fleet"]["k8s"]["ram_gb_per_node"],
         cn["eff_cores_xeon"], cn["k8s_usd"], "Kubernetes infra"),
        ("Qserv node", P["qserv"]["cores_per_node"], P["qserv"]["ram_gb_per_node"],
         P["qserv"]["cores_per_node"], cn["qserv_usd"], "Qserv/Cassandra bare-metal"),
        ("GPU H200", P["current_fleet"]["gpu"]["h200_cores"], P["current_fleet"]["gpu"]["h200_ram_gb"],
         P["current_fleet"]["gpu"]["h200_eff_cores"], cn["gpu_h200_usd"], "GPU compute"),
    ]
    for i, (name, cores, ram, eff, cost, purpose) in enumerate(machines):
        r = 3 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=cores)
        ws.cell(row=r, column=3, value=ram)
        ws.cell(row=r, column=4, value=eff)
        ws.cell(row=r, column=5, value=cost).number_format = num_fmt_dollars()
        ws.cell(row=r, column=6, value=purpose)

    # ── Storage Catalog (rows 10-15) ─────────────────────────────────
    r = 9
    ws.cell(row=r, column=1, value="Storage Catalog").font = Font(bold=True, size=13)
    r = 10
    for ci, hdr in enumerate(["Tier", f"$/TB ({P['general']['scale_year']})"], 1):
        ws.cell(row=r, column=ci, value=hdr)
    apply_header_style(ws, r, 2)

    sc = P["costs"]["storage"]
    tiers = [
        ("Flash (NVMe / Weka)", sc["flash_per_tb"]),
        ("Normal (SATA HDD / Ceph)", sc["normal_per_tb"]),
        ("Latent (Object Store)", sc["latent_per_tb"]),
        ("Tape (LTO)", sc["tape_per_tb"]),
    ]
    for i, (name, pprice) in enumerate(tiers):
        r = 11 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=pprice).number_format = num_fmt_dollars()

    # ── Price Factors & Lifecycle (rows 17-25) ───────────────────────
    r = 17
    ws.cell(row=r, column=1, value="Price & Lifecycle Parameters").font = Font(bold=True, size=13)
    r = 18
    for ci, hdr in enumerate(["Parameter", "Value", "Factor (1-rate)"], 1):
        ws.cell(row=r, column=ci, value=hdr)
    apply_header_style(ws, r, 3)

    pf = P["price_factors"]
    ws.cell(row=19, column=1, value="CPU annual price decrease")
    ws.cell(row=19, column=2, value=pf["cpu_annual_decrease"])
    ws.cell(row=19, column=3).value = f"=1-B19"  # cpuFac

    ws.cell(row=20, column=1, value="Storage annual price decrease")
    ws.cell(row=20, column=2, value=pf["storage_annual_decrease"])
    ws.cell(row=20, column=3).value = f"=1-B20"  # diskFac

    ws.cell(row=21, column=1, value="Qserv annual price decrease")
    ws.cell(row=21, column=2, value=pf["qserv_annual_decrease"])
    ws.cell(row=21, column=3).value = f"=1-B21"  # qservFac

    ws.cell(row=22, column=1, value="Scale year (baseline)")
    ws.cell(row=22, column=2, value=P["general"]["scale_year"])

    ws.cell(row=23, column=1, value="Compute lifetime (years)")
    ws.cell(row=23, column=2, value=P["lifecycle"]["compute_lifetime_years"])

    ws.cell(row=24, column=1, value="Storage lifetime (years)")
    ws.cell(row=24, column=2, value=P["lifecycle"]["storage_lifetime_years"])

    # ── Hosting / Overhead Parameters (rows 27-33) ───────────────────
    r = 27
    ws.cell(row=r, column=1, value="Hosting & Overhead").font = Font(bold=True, size=13)
    h = P["costs"]["hosting"]
    ws.cell(row=28, column=1, value="Nodes per rack")
    ws.cell(row=28, column=2, value=h["nodes_per_rack"])
    ws.cell(row=29, column=1, value="Rack install cost (USD)")
    ws.cell(row=29, column=2, value=h["rack_install_cost"]).number_format = num_fmt_dollars()
    ws.cell(row=30, column=1, value="Hosting per node per year (USD)")
    ws.cell(row=30, column=2, value=h["per_node_per_year"]).number_format = num_fmt_dollars()
    ws.cell(row=31, column=1, value="Overhead % of hardware")
    ws.cell(row=31, column=2, value=h["overhead_pct"])
    ws.cell(row=32, column=1, value="Overhead fixed ($M/year)")
    ws.cell(row=32, column=2, value=h["overhead_fixed_m"])

    # Row references used by Ops Costs:
    # B11=flash $/TB, B12=normal $/TB, B13=latent $/TB, B14=tape $/TB
    # B5=torino cost, B4=torino eff cores (row 4 col 4 for Torino)
    # C19=cpuFac, C20=diskFac, C21=qservFac, B22=scaleYear
    # B23=compute lifetime, B24=storage lifetime
    # B28=nodes/rack, B29=rack cost, B30=hosting/node/yr


# ══════════════════════════════════════════════════════════════════════
# Ops Storage Tab
# ══════════════════════════════════════════════════════════════════════

def build_ops_storage_tab(ws, P, N, FY0):
    ws.sheet_properties.tabColor = "5B9BD5"
    widths = [35, 12] + [16] * N
    set_col_widths(ws, widths)

    obs = P["observing"]
    img = P["imaging"]
    cat = P["catalogs"]
    users = P["users"]
    prompt = P["prompt"]
    qd = P["qserv_data_per_node"]
    rs = P["row_sizes"]

    # ── Section 1: Parameters (rows 1-27) ────────────────────────────
    r = 1
    ws.cell(row=r, column=1, value="Ops Storage — USDF Parameters").font = Font(bold=True, size=13)
    r = 2
    ws.cell(row=r, column=1, value="Parameter")
    ws.cell(row=r, column=2, value="Unit")
    for i in range(N):
        ws.cell(row=r, column=i + 3, value=f"LOY{i+1}")
    apply_header_style(ws, r, N + 2)

    def param_row(row, label, unit, values, fmt=None):
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=unit)
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=i + 3, value=v)
            if fmt:
                c.number_format = fmt

    # Row 3: blank spacer
    # Row 4: Objects
    param_row(4, "Objects", "count", cat["objects_per_year"], num_fmt_int())
    # Row 5: Sources
    param_row(5, "Sources", "count", cat["sources_per_year"], num_fmt_int())
    # Row 6: ForcedSources (formula: multiplier * base)
    ws.cell(row=6, column=1, value="ForcedSources")
    ws.cell(row=6, column=2, value="count")
    for i in range(N):
        base = cat["forced_sources_base_per_year"][i]
        ws.cell(row=6, column=i + 3, value=base).number_format = num_fmt_int()

    # Row 7: blank
    # Row 8: Science users
    param_row(8, "Science users", "users", users["count_per_year"], num_fmt_int())
    # Row 9: Storage per user
    param_row(9, "Storage per science user", "TB", users["tb_per_user_per_year"])

    # Row 10: blank
    # Row 11: LSSTCam image size
    param_row(11, "LSSTCam image size", "TB", [img["lsstcam_image_size_tb"]] * N)
    # Row 12: Raw compression (science + engineering)
    param_row(12, "Raw image compression (science)", "factor", [img["raw_compression"]] * N)
    # Row 13: Lossy compression
    param_row(13, "Lossy image compression", "factor", [img["lossy_compression"]] * N)
    # Row 14: Calibration on-disk vs logical (LSST Key Numbers; see sizing_params)
    param_row(14, "Calibration image compression", "factor", [img["calibration_compression"]] * N)

    # Row 15: Nights per year
    param_row(15, "Observing nights per year", "nights", [obs["nights_per_year"]] * N, num_fmt_int())
    # Row 16: Visits per night
    param_row(16, "Visits per night", "visits", [obs["visits_per_night"]] * N, num_fmt_int())
    # Row 17: Images per visit
    param_row(17, "Images per visit", "images", [obs["images_per_visit"]] * N, num_fmt_int())
    # Row 18: Calibration images per day
    param_row(18, "Calibration images per day", "images", [obs["calibration_images_per_day"]] * N, num_fmt_int())

    # Row 19: Science images per year = nights * visits * images_per_visit
    ws.cell(row=19, column=1, value="LSSTCam Science images (per year)")
    ws.cell(row=19, column=2, value="images")
    for i in range(N):
        c = col(i)
        ws.cell(row=19, column=i + 3).value = f"={c}15*{c}16*{c}17"
        ws.cell(row=19, column=i + 3).number_format = num_fmt_int()

    eng_frac = obs["engineering_fraction"]
    ws.cell(row=20, column=1, value=f"LSSTCam Engineering images ({eng_frac*100:.0f}% of science)")
    ws.cell(row=20, column=2, value="images")
    for i in range(N):
        ws.cell(row=20, column=i + 3).value = f"={eng_frac}*{col(i)}19"
        ws.cell(row=20, column=i + 3).number_format = num_fmt_int()

    # Row 21: Calibration images per year = nights * cals_per_day
    ws.cell(row=21, column=1, value="LSSTCam Calibration images (per year)")
    ws.cell(row=21, column=2, value="images")
    for i in range(N):
        c = col(i)
        ws.cell(row=21, column=i + 3).value = f"={c}15*{c}18"
        ws.cell(row=21, column=i + 3).number_format = num_fmt_int()

    # Row 22: Coadd products
    param_row(22, "Number of coadd data products", "count", [P["precursor"]["coadd_products"]] * N)

    # Row 23-27: Row sizes per LOY
    param_row(23, "Object table row size", "bytes", rs["object_per_year"], num_fmt_int())
    param_row(24, "Object_Extra table row size", "bytes", rs["object_extra_per_year"], num_fmt_int())
    param_row(25, "Source table row size", "bytes", rs["source_per_year"], num_fmt_int())
    param_row(26, "ForcedSource table row size", "bytes", rs["forced_source_per_year"], num_fmt_int())
    param_row(27, "Qserv replication factor", "factor", [P["qserv"]["replication_factor"]] * N)

    # ── Section 2: Dataset Sizing (rows 29-50) ──────────────────────
    r = 29
    ws.cell(row=r, column=1, value="Dataset Sizing (annual production)").font = section_font()
    r = 30
    ws.cell(row=r, column=1, value="Dataset")
    ws.cell(row=r, column=2, value="Unit")
    for i in range(N):
        ws.cell(row=r, column=i + 3, value=f"LOY{i+1}")
    apply_header_style(ws, r, N + 2)

    sf = P["storage_fractions"]
    apdb_ref_tb = sf["apdb_reference_tb"]
    apdb_ref_vis = sf["apdb_reference_visits"]
    ws.cell(row=31, column=1, value="APDB")
    ws.cell(row=31, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=31, column=i + 3).value = f"={apdb_ref_tb}/{apdb_ref_vis}*{c}16*{c}15"
        ws.cell(row=31, column=i + 3).number_format = num_fmt_tb()

    # Row 32: Prompt Tier 1 (accumulates until DR1, then constant)
    ws.cell(row=32, column=1, value="Prompt Tier 1 (cumulative)")
    ws.cell(row=32, column=2, value="TB")
    dr1_year = prompt["dr1_target_year"]
    for i in range(N):
        fy = FY0 + i
        nights = obs["nights_per_year"]
        tb_per_night = prompt["tier1_tb_per_night"]
        if fy <= dr1_year:
            # Accumulates: sum of all prior years + this year
            cumulative = tb_per_night * nights * (i + 1)
        else:
            cumulative = tb_per_night * nights * (dr1_year - FY0 + 1)
        ws.cell(row=32, column=i + 3, value=round(cumulative, 1)).number_format = num_fmt_tb()

    # Row 33: Prompt Tier 2 (30-day rolling)
    ws.cell(row=33, column=1, value="Prompt Tier 2 (30-day rolling)")
    ws.cell(row=33, column=2, value="TB")
    tier2_tb = prompt["tier2_tb_per_night"] * prompt["tier2_retention_days"]
    for i in range(N):
        ws.cell(row=33, column=i + 3, value=round(tier2_tb, 1)).number_format = num_fmt_tb()

    # Row 34: blank
    # Row 35: Raw Images per year (annual production, not cumulative)
    # Science + engineering images use measured raw_compression (Row 12);
    # calibration images use calibration_compression (Row 14) — no measured data.
    ws.cell(row=35, column=1, value="LSSTCam Raw Images (per year)")
    ws.cell(row=35, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=35, column=i + 3).value = (
            f"=({c}19+{c}20)*{c}11*{c}12+{c}21*{c}11*{c}14"
        )
        ws.cell(row=35, column=i + 3).number_format = num_fmt_tb()

    # Row 36: LSSTCam Output Coadd Images (per-DR, constant — area-based)
    # Matches existing spreadsheet: Storage!$C$41/Storage!$C$33 * area * coadd_products
    # = hsc_rc2_output_coadd_tb / hsc_rc2_area * lsstcam_area * 2 ≈ 2700 TB per DR
    prec = P["precursor"]
    coadd_per_dr = (prec["hsc_rc2_output_coadd_tb"] / prec["hsc_rc2_area_deg2"]
                    * prec["lsstcam_area_deg2"] * prec["coadd_products"])
    ws.cell(row=36, column=1, value="LSSTCam Output Coadd Images (per DR)")
    ws.cell(row=36, column=2, value="TB")
    for i in range(N):
        ws.cell(row=36, column=i + 3, value=round(coadd_per_dr, 1)).number_format = num_fmt_tb()

    # Row 37: blank
    # Row 38: LSSTCam Output Images per year (from DRP doc actual per-visit constant)
    ws.cell(row=38, column=1, value="LSSTCam Output Images (per year)")
    ws.cell(row=38, column=2, value="TB")
    out_per_visit = img["output_image_tb_per_visit"]
    for i in range(N):
        c = col(i)
        ws.cell(row=38, column=i + 3).value = f"={out_per_visit}*{c}19/{c}17"
        ws.cell(row=38, column=i + 3).number_format = num_fmt_tb()

    # Row 39: LSSTCam Output Parquet per year (from DRP doc actual per-visit constant)
    ws.cell(row=39, column=1, value="LSSTCam Output Parquet (per year)")
    ws.cell(row=39, column=2, value="TB")
    pq_per_visit = img["parquet_tb_per_visit"]
    for i in range(N):
        c = col(i)
        ws.cell(row=39, column=i + 3).value = f"={pq_per_visit}*{c}19/{c}17"
        ws.cell(row=39, column=i + 3).number_format = num_fmt_tb()

    # Row 40: Sims output (constant)
    param_row(40, "Sims output", "TB", [sf["sims_output_tb"]] * N, num_fmt_tb())

    scratch_frac = sf["scratch_fraction"]
    ws.cell(row=41, column=1, value=f"Scratch ({scratch_frac*100:.0f}% of output images)")
    ws.cell(row=41, column=2, value="TB")
    for i in range(N):
        ws.cell(row=41, column=i + 3).value = f"={scratch_frac}*{col(i)}38"
        ws.cell(row=41, column=i + 3).number_format = num_fmt_tb()

    # Row 42: Qserv Czar/Object = Objects * objectRowSize * replication / 1e12
    ws.cell(row=42, column=1, value="Qserv Czar/Object")
    ws.cell(row=42, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=42, column=i + 3).value = f"={c}4*{c}23*{c}27/1000000000000"
        ws.cell(row=42, column=i + 3).number_format = num_fmt_tb()

    # Row 43: Qserv Database = (Objects*(rowSize+extraRowSize) + Sources*sourceRowSize + ForcedSources*forcedSourceRowSize) * replication / 1e12
    ws.cell(row=43, column=1, value="Qserv Database")
    ws.cell(row=43, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=43, column=i + 3).value = (
            f"=({c}4*({c}23+{c}24)+{c}5*{c}25+{c}6*{c}26)*{c}27/1000000000000"
        )
        ws.cell(row=43, column=i + 3).number_format = num_fmt_tb()

    # Row 44: Science User Home = users * TB/user
    ws.cell(row=44, column=1, value="Science User Home")
    ws.cell(row=44, column=2, value="TB")
    for i in range(N):
        ws.cell(row=44, column=i + 3).value = f"={col(i)}8*{col(i)}9"
        ws.cell(row=44, column=i + 3).number_format = num_fmt_tb()

    misc_frac = sf["misc_overhead_fraction"]
    ws.cell(row=45, column=1, value=f"Other/Misc ({misc_frac*100:.0f}%)")
    ws.cell(row=45, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=45, column=i + 3).value = f"={misc_frac}*SUM({c}31:{c}44)"
        ws.cell(row=45, column=i + 3).number_format = num_fmt_tb()

    # ── Section 3: On-Floor Totals by Tier (rows 47-72) ─────────────
    r = 47
    ws.cell(row=r, column=1, value="USDF Storage On-Floor (by tier)").font = section_font()
    r = 48
    ws.cell(row=r, column=1, value="Storage Tier")
    ws.cell(row=r, column=2, value="Unit")
    for i in range(N):
        ws.cell(row=r, column=i + 3, value=f"LOY{i+1}")
    apply_header_style(ws, r, N + 2)

    # Row 49: Fast — APDB
    ws.cell(row=49, column=1, value="APDB (fast)")
    ws.cell(row=49, column=2, value="TB")
    for i in range(N):
        ws.cell(row=49, column=i + 3).value = f"={col(i)}31"
        ws.cell(row=49, column=i + 3).number_format = num_fmt_tb()

    # Row 50: Fast — Qserv Czar (3-year sliding window)
    ws.cell(row=50, column=1, value="Qserv Czar (fast, 3yr window)")
    ws.cell(row=50, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=50, column=i + 3).value = f"={c}42"
        elif i == 1:
            ws.cell(row=50, column=i + 3).value = f"={col(0)}42+{c}42"
        elif i == 2:
            ws.cell(row=50, column=i + 3).value = f"={col(0)}42+{col(1)}42+{c}42"
        else:
            ws.cell(row=50, column=i + 3).value = f"={col(i-2)}42+{col(i-1)}42+{c}42"
        ws.cell(row=50, column=i + 3).number_format = num_fmt_tb()

    # Row 51: Total Fast = APDB + Qserv Czar + Prompt Tier 2
    ws.cell(row=51, column=1, value="Total Fast (NVMe)")
    ws.cell(row=51, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=51, column=i + 3).value = f"=SUM({c}49:{c}50)+{c}33"
        ws.cell(row=51, column=i + 3).number_format = num_fmt_tb()

    # Row 52: blank
    # Row 53: Normal = sum of output images, parquet, sims, scratch, misc, user home
    ws.cell(row=53, column=1, value="Total Normal (HDD)")
    ws.cell(row=53, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=53, column=i + 3).value = f"=SUM({c}38:{c}45)"
        ws.cell(row=53, column=i + 3).number_format = num_fmt_tb()

    # Row 54: blank
    # Row 55: Qserv Storage (3-year sliding window of Qserv DB)
    ws.cell(row=55, column=1, value="Qserv Storage (3yr window)")
    ws.cell(row=55, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=55, column=i + 3).value = f"={c}43"
        elif i == 1:
            ws.cell(row=55, column=i + 3).value = f"={col(0)}43+{c}43"
        elif i == 2:
            ws.cell(row=55, column=i + 3).value = f"={col(0)}43+{col(1)}43+{c}43"
        else:
            ws.cell(row=55, column=i + 3).value = f"={col(i-2)}43+{col(i-1)}43+{c}43"
        ws.cell(row=55, column=i + 3).number_format = num_fmt_tb()

    # Row 56: blank
    # Rows 57-61: Object Store tier
    # Row 57: Raw Images (cumulative)
    ws.cell(row=57, column=1, value="Raw Images (object store, cumulative)")
    ws.cell(row=57, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=57, column=i + 3).value = f"={c}35"
        else:
            ws.cell(row=57, column=i + 3).value = f"={c}35+{col(i-1)}57"
        ws.cell(row=57, column=i + 3).number_format = num_fmt_tb()

    # Row 58: Output Images lossy (2-year window, lossy compressed)
    ws.cell(row=58, column=1, value="Output Images lossy (obj store, 2yr)")
    ws.cell(row=58, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=58, column=i + 3).value = f"={c}38*{c}13"
        else:
            ws.cell(row=58, column=i + 3).value = f"=({c}38+{col(i-1)}38)*{c}13"
        ws.cell(row=58, column=i + 3).number_format = num_fmt_tb()

    # Row 59: Coadd (3-year window)
    ws.cell(row=59, column=1, value="Coadd Images (obj store, 3yr)")
    ws.cell(row=59, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=59, column=i + 3).value = f"={c}36"
        elif i == 1:
            ws.cell(row=59, column=i + 3).value = f"={col(0)}36+{c}36"
        elif i == 2:
            ws.cell(row=59, column=i + 3).value = f"={col(0)}36+{col(1)}36+{c}36"
        else:
            ws.cell(row=59, column=i + 3).value = f"={col(i-2)}36+{col(i-1)}36+{c}36"
        ws.cell(row=59, column=i + 3).number_format = num_fmt_tb()

    # Row 60: Parquet (3-year window)
    ws.cell(row=60, column=1, value="Output Parquet (obj store, 3yr)")
    ws.cell(row=60, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=60, column=i + 3).value = f"={c}39"
        elif i == 1:
            ws.cell(row=60, column=i + 3).value = f"={col(0)}39+{c}39"
        elif i == 2:
            ws.cell(row=60, column=i + 3).value = f"={col(0)}39+{col(1)}39+{c}39"
        else:
            ws.cell(row=60, column=i + 3).value = f"={col(i-2)}39+{col(i-1)}39+{c}39"
        ws.cell(row=60, column=i + 3).number_format = num_fmt_tb()

    # Row 61: Total Object Store = sum(57:60) + Prompt Tier 1
    ws.cell(row=61, column=1, value="Total Object Store")
    ws.cell(row=61, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=61, column=i + 3).value = f"=SUM({c}57:{c}60)+{c}32"
        ws.cell(row=61, column=i + 3).number_format = num_fmt_tb()

    # Row 62: blank
    # Row 63: Tape — Raw (cumulative, same as row 57)
    ws.cell(row=63, column=1, value="Raw Images (tape, cumulative)")
    ws.cell(row=63, column=2, value="TB")
    for i in range(N):
        ws.cell(row=63, column=i + 3).value = f"={col(i)}57"
        ws.cell(row=63, column=i + 3).number_format = num_fmt_tb()

    # Row 64: Tape — All Data Products/Backup (cumulative normal minus scratch/qserv)
    ws.cell(row=64, column=1, value="Data Products Backup (tape)")
    ws.cell(row=64, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=64, column=i + 3).value = f"={c}53-SUM({c}41:{c}43)"
        else:
            ws.cell(row=64, column=i + 3).value = f"={c}53-SUM({c}41:{c}43)+{col(i-1)}64"
        ws.cell(row=64, column=i + 3).number_format = num_fmt_tb()

    # Row 65: Tape — Object store-only products (cumulative coadd)
    ws.cell(row=65, column=1, value="Object Store Products (tape)")
    ws.cell(row=65, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=65, column=i + 3).value = f"={c}36"
        else:
            ws.cell(row=65, column=i + 3).value = f"={c}36+{col(i-1)}65"
        ws.cell(row=65, column=i + 3).number_format = num_fmt_tb()

    # Row 66: Total Tape = Raw + Backup + Object Store products
    ws.cell(row=66, column=1, value="Total Tape")
    ws.cell(row=66, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=66, column=i + 3).value = f"=SUM({c}63:{c}65)"
        ws.cell(row=66, column=i + 3).number_format = num_fmt_tb()

    # ── Section 4: Grand Total (row 68) ─────────────────────────────
    r = 68
    ws.cell(row=r, column=1, value="Grand Total USDF Storage").font = section_font()
    ws.cell(row=r, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=r, column=i + 3).value = f"={c}51+{c}53+{c}55+{c}61+{c}66"
        ws.cell(row=r, column=i + 3).number_format = num_fmt_tb()


# ══════════════════════════════════════════════════════════════════════
# Ops Compute Tab
# ══════════════════════════════════════════════════════════════════════

def build_ops_compute_tab(ws, P, N, FY0):
    ws.sheet_properties.tabColor = "ED7D31"
    widths = [40, 14] + [18] * N
    set_col_widths(ws, widths)

    drp = P["drp"]
    proc_hours = drp["processing_window_days"] * 24
    idle_mult = float(drp.get("idle_and_inter_stage_multiplier", 1.0))
    dev_pilot = P.get("developer_pilot") or {}
    pilot_node_days_yr = float(dev_pilot.get("additional_node_days_per_year", 0))
    rsp_extra = float(dev_pilot.get("usdf_rsp_extra_fraction_of_drp_concurrent", 0))
    pilot_ch_per_year = pilot_node_days_yr * drp["cores_per_node"] * 24
    drp_core_scale = idle_mult * (1.0 + rsp_extra)

    obs = P["observing"]
    img = P["imaging"]
    qd = P["qserv_data_per_node"]

    # Input data per year of observations (TB)
    input_tb_per_year = (obs["visits_per_night"] * obs["nights_per_year"]
                         * img["lsstcam_image_size_tb"] * img["raw_compression"])
    ch_per_tb = drp["core_hours_per_input_tb"]
    carryover = drp.get("commissioning_carryover_ch", 0)

    lt = P["lifecycle"]["compute_lifetime_years"]
    ap_cores = P["alert_production"]["ap_cores"]
    eff = eff_torino_cores(P)
    fb = fleet_batch(P)
    existing_cores = (fb["milano_nodes"] * fb["milano_cores_per_node"] +
                      fb["torino_nodes"] * fb["torino_cores_per_node"])

    r = 1
    ws.cell(row=r, column=1, value="Ops Compute — USDF").font = Font(bold=True, size=13)
    r = 2
    ws.cell(row=r, column=1, value="Category")
    ws.cell(row=r, column=2, value="Unit")
    for i in range(N):
        ws.cell(row=r, column=i + 3, value=f"LOY{i+1}")
    apply_header_style(ws, r, N + 2)

    # Constant for AP formulas (K8s / prompt — not Slurm batch); off-grid to avoid header clash
    ws.cell(row=99, column=1, value="AP cores (for formulas)").font = Font(italic=True, color="999999")
    ws.cell(row=99, column=2, value=ap_cores).number_format = num_fmt_int()

    # ── DRP (batch) ──────────────────────────────────────────────────
    ws.cell(row=3, column=1, value="DRP core-hours needed")
    ws.cell(row=3, column=2, value="core-hours")
    for i in range(N):
        year_num = i + 1
        ch = input_tb_per_year * year_num * ch_per_tb + year_num * pilot_ch_per_year
        if i == 0:
            ch += carryover
        ws.cell(row=3, column=i + 3, value=round(ch)).number_format = num_fmt_int()

    ws.cell(row=4, column=1, value="DRP cores needed (batch)")
    ws.cell(row=4, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=4, column=i + 3).value = (
            f"=ROUND({c}3/{proc_hours}*{drp_core_scale:.12g},0)"
        )
        ws.cell(row=4, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=5, column=1, value="DRP cores annual increase")
    ws.cell(row=5, column=2, value="cores")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=5, column=i + 3).value = f"={c}4"
        else:
            ws.cell(row=5, column=i + 3).value = f"={c}4-{col(i-1)}4"
        ws.cell(row=5, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=6, column=1, value=f"DRP cores refresh ({lt}yr cycle)")
    ws.cell(row=6, column=2, value="cores")
    for i in range(N):
        if i < lt:
            ws.cell(row=6, column=i + 3, value=0).number_format = num_fmt_int()
        else:
            ws.cell(row=6, column=i + 3).value = f"={col(i-lt)}5"
            ws.cell(row=6, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=7, column=1, value="DRP cores to purchase")
    ws.cell(row=7, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=7, column=i + 3).value = f"={c}5+{c}6"
        ws.cell(row=7, column=i + 3).number_format = num_fmt_int()

    # ── AP (K8s — not batch) ─────────────────────────────────────────
    ws.cell(row=9, column=1, value="AP cores needed (K8s / prompt, not batch)")
    ws.cell(row=9, column=2, value="cores")
    for i in range(N):
        ws.cell(row=9, column=i + 3, value="=$B$99").number_format = num_fmt_int()

    ws.cell(row=10, column=1, value="AP cores initial purchase (LOY1 only)")
    ws.cell(row=10, column=2, value="cores")
    for i in range(N):
        ws.cell(row=10, column=i + 3).value = "=IF(COLUMN()=3,$B$99,0)"
        ws.cell(row=10, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=11, column=1, value=f"AP cores refresh ({lt}yr full replace)")
    ws.cell(row=11, column=2, value="cores")
    for i in range(N):
        ws.cell(row=11, column=i + 3).value = (
            f"=IF(AND(COLUMN()>={3+lt},MOD(COLUMN()-3,{lt})=0),$B$99,0)"
        )
        ws.cell(row=11, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=12, column=1, value="AP cores to purchase (K8s budget)")
    ws.cell(row=12, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=12, column=i + 3).value = f"={c}10+{c}11"
        ws.cell(row=12, column=i + 3).number_format = num_fmt_int()

    # ── DAC / Staff (batch path) ─────────────────────────────────────
    cf = P["compute_fractions"]
    dac_frac = cf["dac_drp_fraction"]
    staff_frac = cf["staff_dac_fraction"]
    ws.cell(row=14, column=1, value=f"DAC/LSP cores ({dac_frac*100:.0f}% of DRP)")
    ws.cell(row=14, column=2, value="cores")
    for i in range(N):
        ws.cell(row=14, column=i + 3).value = f"={dac_frac}*{col(i)}4"
        ws.cell(row=14, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=15, column=1, value="DAC cores annual increase")
    ws.cell(row=15, column=2, value="cores")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=15, column=i + 3).value = f"={c}14"
        else:
            ws.cell(row=15, column=i + 3).value = f"={c}14-{col(i-1)}14"
        ws.cell(row=15, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=16, column=1, value=f"DAC cores refresh ({lt}yr)")
    ws.cell(row=16, column=2, value="cores")
    for i in range(N):
        if i < lt:
            ws.cell(row=16, column=i + 3, value=0).number_format = num_fmt_int()
        else:
            ws.cell(row=16, column=i + 3).value = f"={col(i-lt)}15"
            ws.cell(row=16, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=17, column=1, value="DAC cores to purchase")
    ws.cell(row=17, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=17, column=i + 3).value = f"={c}15+{c}16"
        ws.cell(row=17, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=18, column=1, value=f"Staff LSP cores ({staff_frac*100:.0f}% of DAC)")
    ws.cell(row=18, column=2, value="cores")
    for i in range(N):
        ws.cell(row=18, column=i + 3).value = f"={staff_frac}*{col(i)}14"
        ws.cell(row=18, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=19, column=1, value="Staff cores annual increase")
    ws.cell(row=19, column=2, value="cores")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=19, column=i + 3).value = f"={c}18"
        else:
            ws.cell(row=19, column=i + 3).value = f"={c}18-{col(i-1)}18"
        ws.cell(row=19, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=20, column=1, value=f"Staff cores refresh ({lt}yr)")
    ws.cell(row=20, column=2, value="cores")
    for i in range(N):
        if i < lt:
            ws.cell(row=20, column=i + 3, value=0).number_format = num_fmt_int()
        else:
            ws.cell(row=20, column=i + 3).value = f"={col(i-lt)}19"
            ws.cell(row=20, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=21, column=1, value="Staff cores to purchase")
    ws.cell(row=21, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=21, column=i + 3).value = f"={c}19+{c}20"
        ws.cell(row=21, column=i + 3).number_format = num_fmt_int()

    # ── Qserv (single 3yr window from Ops Storage — no double-count) ──
    ws.cell(row=23, column=1, value="Qserv data per node (capacity this LOY)")
    ws.cell(row=23, column=2, value="TB/node")
    for i in range(N):
        ws.cell(row=23, column=i + 3, value=qd[i]).number_format = num_fmt_tb()

    ws.cell(row=24, column=1, value="Qserv nodes needed (total on floor)")
    ws.cell(row=24, column=2, value="nodes")
    for i in range(N):
        c = col(i)
        ws.cell(row=24, column=i + 3).value = (
            f"=CEILING('Ops Storage'!{c}55/{c}23,1)"
        )
        ws.cell(row=24, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=25, column=1, value="Qserv nodes annual increase")
    ws.cell(row=25, column=2, value="nodes")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=25, column=i + 3).value = f"={c}24"
        else:
            ws.cell(row=25, column=i + 3).value = f"={c}24-{col(i-1)}24"
        ws.cell(row=25, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=26, column=1, value=f"Qserv nodes refresh ({lt}yr)")
    ws.cell(row=26, column=2, value="nodes")
    for i in range(N):
        if i < lt:
            ws.cell(row=26, column=i + 3, value=0).number_format = num_fmt_int()
        else:
            ws.cell(row=26, column=i + 3).value = f"={col(i-lt)}25"
            ws.cell(row=26, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=27, column=1, value="Qserv nodes to purchase")
    ws.cell(row=27, column=2, value="nodes")
    for i in range(N):
        c = col(i)
        ws.cell(row=27, column=i + 3).value = f"={c}25+{c}26"
        ws.cell(row=27, column=i + 3).number_format = num_fmt_int()

    # ── GPU ──────────────────────────────────────────────────────────
    ws.cell(row=29, column=1, value="GPU H200 nodes")
    ws.cell(row=29, column=2, value="nodes")
    gpu = P["current_fleet"]["gpu"]["h200_nodes"]
    for i in range(N):
        ws.cell(row=29, column=i + 3, value=gpu).number_format = num_fmt_int()

    # ── K8s Infrastructure ───────────────────────────────────────────
    k8s_total = P["current_fleet"]["k8s"]["nodes"]
    ws.cell(row=31, column=1, value="K8s infrastructure nodes (workers)")
    ws.cell(row=31, column=2, value="nodes")
    for i in range(N):
        ws.cell(row=31, column=i + 3, value=k8s_total).number_format = num_fmt_int()

    # ── LSP cores per user (DAC only) ───────────────────────────────
    ws.cell(row=33, column=1, value="LSP cores per science user")
    ws.cell(row=33, column=2, value="cores/user")
    for i in range(N):
        ws.cell(row=33, column=i + 3).value = f"={col(i)}14/'Ops Storage'!{col(i)}8"

    # ── Batch totals (excludes AP) ───────────────────────────────────
    ws.cell(row=34, column=1, value="BATCH cores to purchase (DRP+DAC+Staff)").font = section_font()
    ws.cell(row=34, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=34, column=i + 3).value = f"={c}7+{c}17+{c}21"
        ws.cell(row=34, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=35, column=1, value="Batch nodes to purchase (Torino-class)")
    ws.cell(row=35, column=2, value="nodes")
    for i in range(N):
        c = col(i)
        ws.cell(row=35, column=i + 3).value = f"=ROUNDUP({c}34/{eff},0)"
        ws.cell(row=35, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=36, column=1, value="Cumulative batch cores owned")
    ws.cell(row=36, column=2, value="cores")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=36, column=i + 3).value = f"={existing_cores}+{c}34"
        else:
            ws.cell(row=36, column=i + 3).value = f"={col(i-1)}36+{c}34"
        ws.cell(row=36, column=i + 3).number_format = num_fmt_int()


# ══════════════════════════════════════════════════════════════════════
# Ops Costs Tab
# ══════════════════════════════════════════════════════════════════════

def build_ops_costs_tab(ws, P, N, FY0):
    ws.sheet_properties.tabColor = "FF0000"
    widths = [38, 14] + [18] * N + [18]
    set_col_widths(ws, widths)

    total_col = N + 3  # column for "Total" sum

    r = 1
    ws.cell(row=r, column=1, value="Ops Costs — USDF Annual Hardware Spend").font = Font(bold=True, size=13)
    r = 2
    ws.cell(row=r, column=1, value="Category")
    ws.cell(row=r, column=2, value="Unit")
    for i in range(N):
        ws.cell(row=r, column=i + 3, value=f"LOY{i+1} (FY{FY0+i})")
    ws.cell(row=r, column=total_col, value="Total")
    apply_header_style(ws, r, total_col)

    node_cost = P["costs"]["nodes"]["torino_usd"]  # future purchases are Torino
    qserv_cost = P["costs"]["nodes"]["qserv_usd"]
    gpu_cost = P["costs"]["nodes"]["gpu_h200_usd"]
    k8s_cost = P["costs"]["nodes"]["k8s_usd"]

    flash_tb = P["costs"]["storage"]["flash_per_tb"]
    normal_tb = P["costs"]["storage"]["normal_per_tb"]
    latent_tb = P["costs"]["storage"]["latent_per_tb"]
    tape_tb = P["costs"]["storage"]["tape_per_tb"]

    overhead_pct = P["costs"]["hosting"]["overhead_pct"]
    overhead_fixed = P["costs"]["hosting"]["overhead_fixed_m"]
    hosting_per_node = P["costs"]["hosting"]["per_node_per_year"]

    cpu_fac = 1 - P["price_factors"]["cpu_annual_decrease"]
    disk_fac = 1 - P["price_factors"]["storage_annual_decrease"]
    qserv_fac = 1 - P["price_factors"]["qserv_annual_decrease"]
    scale_yr = P["general"]["scale_year"]

    def total_formula(row_num):
        first = col(0)
        last = col(N - 1)
        return f"=SUM({first}{row_num}:{last}{row_num})"

    # ── Compute Costs (rows 3-10) ────────────────────────────────────
    # Row 3: Batch compute cost (at scale_year pricing)
    ws.cell(row=3, column=1, value="Batch compute (nodes * Torino price)")
    ws.cell(row=3, column=2, value="$M")
    for i in range(N):
        c = col(i)
        ws.cell(row=3, column=i + 3).value = f"='Ops Compute'!{c}35*{node_cost}/1000000"
        ws.cell(row=3, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=3, column=total_col).value = total_formula(3)
    ws.cell(row=3, column=total_col).number_format = num_fmt_millions()

    # Row 4: Batch compute (price-adjusted)
    ws.cell(row=4, column=1, value="Batch compute (price-adjusted)")
    ws.cell(row=4, column=2, value="$M")
    for i in range(N):
        c = col(i)
        fy = FY0 + i
        ws.cell(row=4, column=i + 3).value = f"={c}3*POWER({cpu_fac},{fy}-{scale_yr})"
        ws.cell(row=4, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=4, column=total_col).value = total_formula(4)
    ws.cell(row=4, column=total_col).number_format = num_fmt_millions()

    # Row 5: Qserv compute cost
    ws.cell(row=5, column=1, value="Qserv node purchases")
    ws.cell(row=5, column=2, value="$M")
    for i in range(N):
        c = col(i)
        ws.cell(row=5, column=i + 3).value = (
            f"='Ops Compute'!{c}27*{qserv_cost}/1000000"
        )
        ws.cell(row=5, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=5, column=total_col).value = total_formula(5)
    ws.cell(row=5, column=total_col).number_format = num_fmt_millions()

    # Row 6: Qserv (price-adjusted)
    ws.cell(row=6, column=1, value="Qserv (price-adjusted)")
    ws.cell(row=6, column=2, value="$M")
    for i in range(N):
        c = col(i)
        fy = FY0 + i
        ws.cell(row=6, column=i + 3).value = f"={c}5*POWER({qserv_fac},{fy}-{scale_yr})"
        ws.cell(row=6, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=6, column=total_col).value = total_formula(6)
    ws.cell(row=6, column=total_col).number_format = num_fmt_millions()

    # Row 7: GPU cost
    ws.cell(row=7, column=1, value="GPU H200 purchases")
    ws.cell(row=7, column=2, value="$M")
    for i in range(N):
        c = col(i)
        ws.cell(row=7, column=i + 3).value = f"='Ops Compute'!{c}29*{gpu_cost}/1000000"
        ws.cell(row=7, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=7, column=total_col).value = total_formula(7)
    ws.cell(row=7, column=total_col).number_format = num_fmt_millions()

    # Row 8: K8s cost
    ws.cell(row=8, column=1, value="K8s infrastructure")
    ws.cell(row=8, column=2, value="$M")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=8, column=i + 3).value = f"='Ops Compute'!{c}31*{k8s_cost}/1000000"
        else:
            ws.cell(row=8, column=i + 3, value=0).number_format = num_fmt_millions()
    ws.cell(row=8, column=total_col).value = total_formula(8)
    ws.cell(row=8, column=total_col).number_format = num_fmt_millions()

    # Row 9: Total compute cost
    ws.cell(row=9, column=1, value="Total Compute Cost").font = section_font()
    ws.cell(row=9, column=2, value="$M")
    for i in range(N):
        c = col(i)
        ws.cell(row=9, column=i + 3).value = f"={c}4+{c}6+{c}7+{c}8"
        ws.cell(row=9, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=9, column=total_col).value = total_formula(9)
    ws.cell(row=9, column=total_col).number_format = num_fmt_millions()

    # ── Storage Costs (rows 11-18) ───────────────────────────────────
    r = 11
    ws.cell(row=r, column=1, value="Storage Costs").font = section_font()

    # Row 12: Fast storage cost
    ws.cell(row=12, column=1, value="Flash storage (NVMe)")
    ws.cell(row=12, column=2, value="$M")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=12, column=i + 3).value = f"='Ops Storage'!{c}51*{flash_tb}/1000000"
        else:
            ws.cell(row=12, column=i + 3).value = (
                f"=MAX(0,'Ops Storage'!{c}51-'Ops Storage'!{col(i-1)}51)*{flash_tb}/1000000"
            )
        ws.cell(row=12, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=12, column=total_col).value = total_formula(12)
    ws.cell(row=12, column=total_col).number_format = num_fmt_millions()

    # Row 13: Normal storage cost
    ws.cell(row=13, column=1, value="Normal storage (HDD)")
    ws.cell(row=13, column=2, value="$M")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=13, column=i + 3).value = f"='Ops Storage'!{c}53*{normal_tb}/1000000"
        else:
            ws.cell(row=13, column=i + 3).value = (
                f"=MAX(0,'Ops Storage'!{c}53-'Ops Storage'!{col(i-1)}53)*{normal_tb}/1000000"
            )
        ws.cell(row=13, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=13, column=total_col).value = total_formula(13)
    ws.cell(row=13, column=total_col).number_format = num_fmt_millions()

    # Row 14: Object store cost
    ws.cell(row=14, column=1, value="Object Store (latent)")
    ws.cell(row=14, column=2, value="$M")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=14, column=i + 3).value = f"='Ops Storage'!{c}61*{latent_tb}/1000000"
        else:
            ws.cell(row=14, column=i + 3).value = (
                f"=MAX(0,'Ops Storage'!{c}61-'Ops Storage'!{col(i-1)}61)*{latent_tb}/1000000"
            )
        ws.cell(row=14, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=14, column=total_col).value = total_formula(14)
    ws.cell(row=14, column=total_col).number_format = num_fmt_millions()

    # Row 15: Tape cost
    ws.cell(row=15, column=1, value="Tape storage")
    ws.cell(row=15, column=2, value="$M")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=15, column=i + 3).value = f"='Ops Storage'!{c}66*{tape_tb}/1000000"
        else:
            ws.cell(row=15, column=i + 3).value = (
                f"=MAX(0,'Ops Storage'!{c}66-'Ops Storage'!{col(i-1)}66)*{tape_tb}/1000000"
            )
        ws.cell(row=15, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=15, column=total_col).value = total_formula(15)
    ws.cell(row=15, column=total_col).number_format = num_fmt_millions()

    # Row 16: Storage price-adjusted
    ws.cell(row=16, column=1, value="Total Storage (price-adjusted)")
    ws.cell(row=16, column=2, value="$M")
    for i in range(N):
        c = col(i)
        fy = FY0 + i
        ws.cell(row=16, column=i + 3).value = (
            f"=({c}12+{c}13+{c}14+{c}15)*POWER({disk_fac},{fy}-{scale_yr})"
        )
        ws.cell(row=16, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=16, column=total_col).value = total_formula(16)
    ws.cell(row=16, column=total_col).number_format = num_fmt_millions()

    # Row 17: Total storage at base pricing
    ws.cell(row=17, column=1, value="Total Storage (base pricing)").font = section_font()
    ws.cell(row=17, column=2, value="$M")
    for i in range(N):
        c = col(i)
        ws.cell(row=17, column=i + 3).value = f"={c}12+{c}13+{c}14+{c}15"
        ws.cell(row=17, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=17, column=total_col).value = total_formula(17)
    ws.cell(row=17, column=total_col).number_format = num_fmt_millions()

    # ── Hosting / Overhead (rows 19-22) ──────────────────────────────
    r = 19
    ws.cell(row=r, column=1, value="Hosting & Overhead").font = section_font()

    # Row 20: Hosting cost = CEILING(cumulative_cores / eff_cores_per_node) * hosting_per_node / 1e6
    ws.cell(row=20, column=1, value="USDF hosting (nodes * $/node/yr)")
    ws.cell(row=20, column=2, value="$M")
    for i in range(N):
        c = col(i)
        ws.cell(row=20, column=i + 3).value = (
            f"=CEILING('Ops Compute'!{c}36/'Model'!$D$4,1)*'Model'!$B$30/1000000"
        )
        ws.cell(row=20, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=20, column=total_col).value = total_formula(20)
    ws.cell(row=20, column=total_col).number_format = num_fmt_millions()

    # Row 21: Overhead = 14% of hardware + $150K fixed
    ws.cell(row=21, column=1, value=f"Overhead ({overhead_pct*100:.0f}% HW + ${overhead_fixed}M fixed)")
    ws.cell(row=21, column=2, value="$M")
    for i in range(N):
        c = col(i)
        ws.cell(row=21, column=i + 3).value = (
            f"={overhead_pct}*({c}9+{c}17)+{overhead_fixed}"
        )
        ws.cell(row=21, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=21, column=total_col).value = total_formula(21)
    ws.cell(row=21, column=total_col).number_format = num_fmt_millions()

    # ── Grand Totals (rows 23-26) ────────────────────────────────────
    r = 23
    ws.cell(row=r, column=1, value="GRAND TOTALS").font = Font(bold=True, size=13, color="C00000")

    # Row 24: Annual total (base pricing)
    ws.cell(row=24, column=1, value="Annual Total (base pricing)")
    ws.cell(row=24, column=2, value="$M")
    for i in range(N):
        c = col(i)
        ws.cell(row=24, column=i + 3).value = f"={c}9+{c}17+{c}20+{c}21"
        ws.cell(row=24, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=24, column=total_col).value = total_formula(24)
    ws.cell(row=24, column=total_col).number_format = num_fmt_millions()

    # Row 25: Annual total (price-adjusted)
    ws.cell(row=25, column=1, value="Annual Total (price-adjusted)")
    ws.cell(row=25, column=2, value="$M")
    for i in range(N):
        c = col(i)
        ws.cell(row=25, column=i + 3).value = f"={c}4+{c}6+{c}7+{c}8+{c}16+{c}20+{c}21"
        ws.cell(row=25, column=i + 3).number_format = num_fmt_millions()
    ws.cell(row=25, column=total_col).value = total_formula(25)
    ws.cell(row=25, column=total_col).number_format = num_fmt_millions()

    # Row 26: Cumulative (price-adjusted)
    ws.cell(row=26, column=1, value="Cumulative (price-adjusted)")
    ws.cell(row=26, column=2, value="$M")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=26, column=i + 3).value = f"={c}25"
        else:
            ws.cell(row=26, column=i + 3).value = f"={col(i-1)}26+{c}25"
        ws.cell(row=26, column=i + 3).number_format = num_fmt_millions()


    # DR Readiness has been replaced by build_yearly_readiness_tab + build_purchase_plan_tab


# ══════════════════════════════════════════════════════════════════════
# Hardware Refresh (EOL) Tab
# ══════════════════════════════════════════════════════════════════════

def build_hardware_refresh_tab(ws, P, N, FY0):
    """Cohort-based replacements: FY hits EOL when FY > delivery_year and
    MOD(FY - delivery_year, lifetime_years) = 0."""
    ws.sheet_properties.tabColor = "548235"
    widths = [36, 14, 10, 12, 10, 14] + [11] * N
    set_col_widths(ws, widths)

    hdr = P.get("hardware_refresh", {})
    cohorts = hdr.get("cohorts", [])

    ws.cell(row=1, column=1, value="Hardware Refresh / EOL (annual units to replace)").font = Font(
        bold=True, size=14
    )
    ws.cell(row=2, column=1,
            value="Each cohort: replace `Qty` in FY when FY > delivery_year and "
                  "MOD(FY − delivery_year, lifetime_years) = 0. "
                  "Lifetimes: compute-class 3 yr, JBOD/NVMe 5 yr (from YAML). "
                  "Source: Rubin Hardware Summary for March 2025 JTM + sizing_params.yaml."
            ).font = Font(italic=True, color="444444")

    base_col = 7  # column G = first LOY FY
    hdr_row = 5
    ws.cell(row=hdr_row, column=1, value="Fiscal Year (FY)").font = Font(bold=True)
    for i in range(N):
        c = get_column_letter(base_col + i)
        fy = FY0 + i
        ws.cell(row=hdr_row, column=base_col + i, value=fy).number_format = num_fmt_int()

    hrow = 6
    ws.cell(row=hrow, column=1, value="Label")
    ws.cell(row=hrow, column=2, value="Category")
    ws.cell(row=hrow, column=3, value="Qty")
    ws.cell(row=hrow, column=4, value="Delivery FY")
    ws.cell(row=hrow, column=5, value="Lifetime (yr)")
    ws.cell(row=hrow, column=6, value="First EOL FY")
    for i in range(N):
        ws.cell(row=hrow, column=base_col + i, value=f"LOY{i+1}")
    apply_header_style(ws, hrow, base_col + N - 1)

    for idx, coh in enumerate(cohorts):
        rr = 7 + idx
        ws.cell(row=rr, column=1, value=coh.get("label", ""))
        ws.cell(row=rr, column=2, value=coh.get("category", ""))
        ws.cell(row=rr, column=3, value=coh.get("quantity", 0)).number_format = num_fmt_int()
        d_y = coh.get("delivery_year", 0)
        lt_y = int(coh.get("lifetime_years", 3))
        ws.cell(row=rr, column=4, value=d_y).number_format = num_fmt_int()
        ws.cell(row=rr, column=5, value=lt_y).number_format = num_fmt_int()
        ws.cell(row=rr, column=6, value=d_y + lt_y).number_format = num_fmt_int()
        for i in range(N):
            cc = get_column_letter(base_col + i)
            ws.cell(row=rr, column=base_col + i).value = (
                f"=IF(AND({cc}${hdr_row}>$D{rr},MOD({cc}${hdr_row}-$D{rr},$E{rr})=0),$C{rr},0)"
            )
            ws.cell(row=rr, column=base_col + i).number_format = num_fmt_int()

    tot_row = 7 + len(cohorts)
    ws.cell(row=tot_row, column=1, value="TOTAL units to replace (all cohorts)").font = Font(bold=True)
    for i in range(N):
        cc = get_column_letter(base_col + i)
        first_data = 7
        last_data = 6 + len(cohorts)
        if last_data >= first_data:
            ws.cell(row=tot_row, column=base_col + i).value = (
                f"=SUM({cc}{first_data}:{cc}{last_data})"
            )
        else:
            ws.cell(row=tot_row, column=base_col + i, value=0)
        ws.cell(row=tot_row, column=base_col + i).number_format = num_fmt_int()

    # Optional subtotals by category (static note — user can pivot from data)
    note_r = tot_row + 2
    ws.cell(row=note_r, column=1,
            value="Note: Milan batch / Qserv / k8s / Cassandra use compute lifetime; "
                  "JBOD / NVMe use storage lifetime. Edit cohorts in sizing_params.yaml."
            ).font = Font(italic=True, color="666666")


# ══════════════════════════════════════════════════════════════════════
# Charts Tab
# ══════════════════════════════════════════════════════════════════════

def build_charts_tab(ws, P, N, FY0):
    ws.sheet_properties.tabColor = "00B050"
    set_col_widths(ws, [18] + [14] * 18)

    pb_fmt = "#,##0.0"
    ch_fmt = "#,##0"

    # ── Data Block 1 (row 1): Storage by tier in PB ──────────────────
    h1 = ["Year", "Fast (PB)", "Normal (PB)", "Qserv (PB)",
           "Object Store (PB)", "Tape (PB)"]
    for ci, h in enumerate(h1, 1):
        ws.cell(row=1, column=ci, value=h)
    apply_header_style(ws, 1, len(h1))

    for i in range(N):
        r = i + 2
        c = col(i)
        ws.cell(row=r, column=1, value=f"LOY{i+1} (FY{FY0+i})")
        ws.cell(row=r, column=2).value = f"='Ops Storage'!{c}51/1000"
        ws.cell(row=r, column=2).number_format = pb_fmt
        ws.cell(row=r, column=3).value = f"='Ops Storage'!{c}53/1000"
        ws.cell(row=r, column=3).number_format = pb_fmt
        ws.cell(row=r, column=4).value = f"='Ops Storage'!{c}55/1000"
        ws.cell(row=r, column=4).number_format = pb_fmt
        ws.cell(row=r, column=5).value = f"='Ops Storage'!{c}61/1000"
        ws.cell(row=r, column=5).number_format = pb_fmt
        ws.cell(row=r, column=6).value = f"='Ops Storage'!{c}66/1000"
        ws.cell(row=r, column=6).number_format = pb_fmt

    last_r1 = N + 1

    # ── Data Block 2 (row N+3): Detail tiers (Flash, Qserv, ObjStore) ─
    detail_start = last_r1 + 2
    h2 = ["Year", "Flash (PB)", "Qserv (PB)", "Object Store (PB)"]
    for ci, h in enumerate(h2, 1):
        ws.cell(row=detail_start, column=ci, value=h)
    apply_header_style(ws, detail_start, len(h2))

    for i in range(N):
        r = detail_start + 1 + i
        c = col(i)
        ws.cell(row=r, column=1, value=f"LOY{i+1}")
        ws.cell(row=r, column=2).value = f"='Ops Storage'!{c}51/1000"
        ws.cell(row=r, column=2).number_format = pb_fmt
        ws.cell(row=r, column=3).value = f"='Ops Storage'!{c}55/1000"
        ws.cell(row=r, column=3).number_format = pb_fmt
        ws.cell(row=r, column=4).value = f"='Ops Storage'!{c}61/1000"
        ws.cell(row=r, column=4).number_format = pb_fmt

    last_r2 = detail_start + N

    # ── Data Block 3 (row ...): Data product breakdown in PB ──────────
    dp_start = last_r2 + 2
    h3 = ["Year", "Raw Images (PB)", "Processed Images (PB)",
           "Co-added Images (PB)", "Parquet Tables (PB)",
           "Catalog Database (PB)"]
    for ci, h in enumerate(h3, 1):
        ws.cell(row=dp_start, column=ci, value=h)
    apply_header_style(ws, dp_start, len(h3))

    for i in range(N):
        r = dp_start + 1 + i
        c = col(i)
        ws.cell(row=r, column=1, value=f"LOY{i+1}")
        ws.cell(row=r, column=2).value = f"='Ops Storage'!{c}57/1000"
        ws.cell(row=r, column=2).number_format = pb_fmt
        ws.cell(row=r, column=3).value = f"='Ops Storage'!{c}58/1000"
        ws.cell(row=r, column=3).number_format = pb_fmt
        ws.cell(row=r, column=4).value = f"='Ops Storage'!{c}59/1000"
        ws.cell(row=r, column=4).number_format = pb_fmt
        ws.cell(row=r, column=5).value = f"='Ops Storage'!{c}60/1000"
        ws.cell(row=r, column=5).number_format = pb_fmt
        ws.cell(row=r, column=6).value = f"=('Ops Storage'!{c}42+'Ops Storage'!{c}43)/1000"
        ws.cell(row=r, column=6).number_format = pb_fmt

    last_r3 = dp_start + N

    # ── Data Block 4 (row ...): DRP Core-Hours ────────────────────────
    ch_start = last_r3 + 2
    h4 = ["Year", "DRP Core-Hours"]
    for ci, h in enumerate(h4, 1):
        ws.cell(row=ch_start, column=ci, value=h)
    apply_header_style(ws, ch_start, len(h4))

    for i in range(N):
        r = ch_start + 1 + i
        ws.cell(row=r, column=1, value=f"LOY{i+1}")
        ws.cell(row=r, column=2).value = f"='Ops Compute'!{col(i)}3"
        ws.cell(row=r, column=2).number_format = ch_fmt

    last_r4 = ch_start + N

    # ── Data Block 5: Cost ────────────────────────────────────────────
    co_start = last_r4 + 2
    h5 = ["Year", "Annual Cost ($M)", "Cumulative Cost ($M)"]
    for ci, h in enumerate(h5, 1):
        ws.cell(row=co_start, column=ci, value=h)
    apply_header_style(ws, co_start, len(h5))

    for i in range(N):
        r = co_start + 1 + i
        ws.cell(row=r, column=1, value=f"LOY{i+1}")
        ws.cell(row=r, column=2).value = f"='Ops Costs'!{col(i)}25"
        ws.cell(row=r, column=2).number_format = num_fmt_millions()
        ws.cell(row=r, column=3).value = f"='Ops Costs'!{col(i)}26"
        ws.cell(row=r, column=3).number_format = num_fmt_millions()

    last_r5 = co_start + N

    # ══════════════════════════════════════════════════════════════════
    # CHARTS
    # ══════════════════════════════════════════════════════════════════

    chart_col = 8  # Place charts starting at column H

    # Chart 1: USDF Storage Forecast by Tier (line chart, PB)
    c1 = LineChart()
    c1.title = "USDF Storage Forecast by Tier"
    c1.y_axis.title = "Petabytes"
    c1.x_axis.title = "Year"
    c1.style = 10
    c1.width = 28
    c1.height = 16
    cats1 = Reference(ws, min_col=1, min_row=2, max_row=last_r1)
    for ci in range(2, 7):
        data = Reference(ws, min_col=ci, min_row=1, max_row=last_r1)
        c1.add_data(data, titles_from_data=True)
    c1.set_categories(cats1)
    ws.add_chart(c1, f"{get_column_letter(chart_col)}1")

    # Chart 2: Detail tiers (Flash, Qserv, Object Store — PB)
    c2 = LineChart()
    c2.title = "Storage Detail: Flash / Qserv / Object Store"
    c2.y_axis.title = "Petabytes"
    c2.x_axis.title = "Year"
    c2.style = 10
    c2.width = 28
    c2.height = 16
    cats2 = Reference(ws, min_col=1, min_row=detail_start + 1,
                       max_row=last_r2)
    for ci in range(2, 5):
        data = Reference(ws, min_col=ci, min_row=detail_start,
                          max_row=last_r2)
        c2.add_data(data, titles_from_data=True)
    c2.set_categories(cats2)
    ws.add_chart(c2, f"{get_column_letter(chart_col)}17")

    # Chart 3: Data product breakdown (stacked area, PB)
    c3 = AreaChart()
    c3.title = "Data Product Size Breakdown"
    c3.y_axis.title = "Petabytes"
    c3.x_axis.title = "Year"
    c3.style = 10
    c3.width = 28
    c3.height = 16
    c3.grouping = "stacked"
    cats3 = Reference(ws, min_col=1, min_row=dp_start + 1,
                       max_row=last_r3)
    for ci in range(2, 7):
        data = Reference(ws, min_col=ci, min_row=dp_start,
                          max_row=last_r3)
        c3.add_data(data, titles_from_data=True)
    c3.set_categories(cats3)
    ws.add_chart(c3, f"{get_column_letter(chart_col)}33")

    # Chart 4: DRP Core-Hours (line)
    c4 = LineChart()
    c4.title = "DRP Core-Hours Projection"
    c4.y_axis.title = "Core-Hours"
    c4.x_axis.title = "Year"
    c4.style = 10
    c4.width = 28
    c4.height = 16
    cats4 = Reference(ws, min_col=1, min_row=ch_start + 1,
                       max_row=last_r4)
    data4 = Reference(ws, min_col=2, min_row=ch_start,
                       max_row=last_r4)
    c4.add_data(data4, titles_from_data=True)
    c4.set_categories(cats4)
    ws.add_chart(c4, f"{get_column_letter(chart_col)}49")

    # Chart 5: Cost Forecast (line)
    c5 = LineChart()
    c5.title = "USDF Cost Forecast"
    c5.y_axis.title = "$M"
    c5.x_axis.title = "Year"
    c5.style = 10
    c5.width = 28
    c5.height = 16
    cats5 = Reference(ws, min_col=1, min_row=co_start + 1,
                       max_row=last_r5)
    for ci in [2, 3]:
        data = Reference(ws, min_col=ci, min_row=co_start,
                          max_row=last_r5)
        c5.add_data(data, titles_from_data=True)
    c5.set_categories(cats5)
    ws.add_chart(c5, f"{get_column_letter(chart_col)}65")


# ══════════════════════════════════════════════════════════════════════
# Purchase Plan Tab
# ══════════════════════════════════════════════════════════════════════

def build_purchase_plan_tab(ws, P, N, FY0):
    ws.sheet_properties.tabColor = "002060"
    widths = [42, 14] + [18] * N
    set_col_widths(ws, widths)

    milestones = P.get("milestones_by_loy", {})
    installed_tb = P["current_fleet"]["storage"]["total_installed_pb"] * 1000
    fleet = P["current_fleet"]["batch"]
    existing_cores = (fleet["milano_nodes"] * fleet["milano_cores_per_node"] +
                      fleet["torino_nodes"] * fleet["torino_cores_per_node"])
    eff = eff_torino_cores(P)
    lt = P["lifecycle"]["compute_lifetime_years"]
    k8s_base = P["current_fleet"]["k8s"]["nodes"]
    k8s_growth = P["k8s"]["growth_rate_per_year"]
    dp2_frac = P["dp2"]["drp_fraction"]
    cur = P.get("current_year", {})
    months_remaining = cur["months_remaining"]

    r = 1
    ws.cell(row=r, column=1, value="Annual Purchase Plan — USDF").font = Font(bold=True, size=14)
    r = 2
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Unit")
    for i in range(N):
        ws.cell(row=r, column=i + 3, value=f"LOY{i+1}")
    apply_header_style(ws, r, N + 2)

    # Row 3: FY
    ws.cell(row=3, column=1, value="Fiscal Year")
    for i in range(N):
        ws.cell(row=3, column=i + 3, value=f"FY{FY0+i}")

    # Row 4: Milestone
    ws.cell(row=4, column=1, value="Milestone")
    for i in range(N):
        ms = milestones.get(i + 1, "")
        c = ws.cell(row=4, column=i + 3, value=ms)
        if ms:
            c.font = Font(bold=True, color="C00000")

    # ── COMPUTE ──────────────────────────────────────────────────────
    ws.cell(row=6, column=1, value="COMPUTE PURCHASES").font = section_font()

    ws.cell(row=7, column=1, value="DRP cores needed")
    ws.cell(row=7, column=2, value="cores")
    for i in range(N):
        ws.cell(row=7, column=i + 3).value = f"='Ops Compute'!{col(i)}4"
        ws.cell(row=7, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=8, column=1, value="USDF DRP cores (100%)")
    ws.cell(row=8, column=2, value="cores")
    for i in range(N):
        ws.cell(row=8, column=i + 3).value = f"={col(i)}7"
        ws.cell(row=8, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=9, column=1, value="AP cores")
    ws.cell(row=9, column=2, value="cores")
    for i in range(N):
        ws.cell(row=9, column=i + 3).value = f"='Ops Compute'!{col(i)}9"
        ws.cell(row=9, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=10, column=1, value="DAC/LSP + Staff cores")
    ws.cell(row=10, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=10, column=i + 3).value = f"='Ops Compute'!{c}14+'Ops Compute'!{c}18"
        ws.cell(row=10, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=11, column=1, value="Total batch cores needed (DRP+DAC+Staff)")
    ws.cell(row=11, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=11, column=i + 3).value = f"={c}8+{c}10"
        ws.cell(row=11, column=i + 3).number_format = num_fmt_int()
    ws.cell(row=11, column=1).font = Font(bold=True)

    ws.cell(row=12, column=1, value="Cumulative cores available")
    ws.cell(row=12, column=2, value="cores")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=12, column=i + 3).value = f"={existing_cores}+MAX(0,{c}14)"
        else:
            ws.cell(row=12, column=i + 3).value = f"={col(i-1)}12+MAX(0,{c}14)"
        ws.cell(row=12, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=13, column=1, value="Compute gap (need − available)")
    ws.cell(row=13, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=13, column=i + 3).value = f"={c}11-{c}12"
        ws.cell(row=13, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=14, column=1, value="Batch cores to purchase (= Ops Compute)")
    ws.cell(row=14, column=2, value="cores")
    for i in range(N):
        c = col(i)
        ws.cell(row=14, column=i + 3).value = f"='Ops Compute'!{c}34"
        ws.cell(row=14, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=15, column=1, value="Batch nodes to purchase (= Ops Compute)")
    ws.cell(row=15, column=2, value="nodes")
    for i in range(N):
        c = col(i)
        ws.cell(row=15, column=i + 3).value = f"='Ops Compute'!{c}35"
        ws.cell(row=15, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=16, column=1, value="K8s nodes needed")
    ws.cell(row=16, column=2, value="nodes")
    for i in range(N):
        ws.cell(row=16, column=i + 3, value=round(k8s_base * (1 + k8s_growth) ** i)).number_format = num_fmt_int()

    ws.cell(row=17, column=1, value="K8s incremental purchase")
    ws.cell(row=17, column=2, value="nodes")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=17, column=i + 3, value=0).number_format = num_fmt_int()
        else:
            ws.cell(row=17, column=i + 3).value = f"=MAX(0,{c}16-{col(i-1)}16)"
            ws.cell(row=17, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=18, column=1, value="Qserv nodes to purchase (annual, = Ops Compute)")
    ws.cell(row=18, column=2, value="nodes")
    for i in range(N):
        c = col(i)
        ws.cell(row=18, column=i + 3).value = f"='Ops Compute'!{c}27"
        ws.cell(row=18, column=i + 3).number_format = num_fmt_int()

    # ── STORAGE ──────────────────────────────────────────────────────
    ws.cell(row=20, column=1, value="STORAGE PURCHASES (by tier)").font = section_font()

    tiers = [
        (21, 22, "Flash (NVMe)", "TB", "'Ops Storage'!{c}51"),
        (23, 24, "Normal HDD", "TB", "'Ops Storage'!{c}53"),
        (25, 26, "Object Store", "TB", "'Ops Storage'!{c}61"),
        (27, 28, "Qserv", "TB", "'Ops Storage'!{c}55"),
    ]
    for floor_r, incr_r, label, unit, ref_tmpl in tiers:
        ws.cell(row=floor_r, column=1, value=f"{label} on floor")
        ws.cell(row=floor_r, column=2, value=unit)
        ws.cell(row=incr_r, column=1, value=f"{label} incremental purchase")
        ws.cell(row=incr_r, column=2, value=unit)
        for i in range(N):
            c = col(i)
            ws.cell(row=floor_r, column=i + 3).value = f"={ref_tmpl.format(c=c)}"
            ws.cell(row=floor_r, column=i + 3).number_format = num_fmt_tb()
            if i == 0:
                ws.cell(row=incr_r, column=i + 3).value = f"={c}{floor_r}"
            else:
                ws.cell(row=incr_r, column=i + 3).value = f"=MAX(0,{c}{floor_r}-{col(i-1)}{floor_r})"
            ws.cell(row=incr_r, column=i + 3).number_format = num_fmt_tb()

    # Row 30: Online (hot+warm) total on floor
    ws.cell(row=30, column=1, value="ONLINE total on floor (hot+warm)")
    ws.cell(row=30, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=30, column=i + 3).value = f"={c}21+{c}23+{c}25+{c}27"
        ws.cell(row=30, column=i + 3).number_format = num_fmt_tb()
    ws.cell(row=30, column=1).font = Font(bold=True)

    # Row 31: Online cumulative available (starts from installed)
    ws.cell(row=31, column=1, value="Online available (installed + purchases)")
    ws.cell(row=31, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=31, column=i + 3).value = f"={installed_tb}+{c}32"
        else:
            ws.cell(row=31, column=i + 3).value = f"={col(i-1)}31+{c}32"
        ws.cell(row=31, column=i + 3).number_format = num_fmt_tb()

    # Row 32: Online incremental purchase
    ws.cell(row=32, column=1, value="Online purchase (hot+warm)")
    ws.cell(row=32, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=32, column=i + 3).value = f"=MAX(0,{c}30-{installed_tb})"
        else:
            ws.cell(row=32, column=i + 3).value = f"=MAX(0,{c}30-{col(i-1)}31)"
        ws.cell(row=32, column=i + 3).number_format = num_fmt_tb()
    ws.cell(row=32, column=1).font = Font(bold=True)

    # Row 33: Online purchase in PB
    ws.cell(row=33, column=1, value="Online purchase (PB)")
    ws.cell(row=33, column=2, value="PB")
    for i in range(N):
        ws.cell(row=33, column=i + 3).value = f"={col(i)}32/1000"
        ws.cell(row=33, column=i + 3).number_format = "#,##0.0"

    # Row 35: Tape section
    ws.cell(row=35, column=1, value="Tape on floor")
    ws.cell(row=35, column=2, value="TB")
    for i in range(N):
        ws.cell(row=35, column=i + 3).value = f"='Ops Storage'!{col(i)}66"
        ws.cell(row=35, column=i + 3).number_format = num_fmt_tb()

    ws.cell(row=36, column=1, value="Tape incremental purchase")
    ws.cell(row=36, column=2, value="TB")
    for i in range(N):
        c = col(i)
        if i == 0:
            ws.cell(row=36, column=i + 3).value = f"={c}35"
        else:
            ws.cell(row=36, column=i + 3).value = f"=MAX(0,{c}35-{col(i-1)}35)"
        ws.cell(row=36, column=i + 3).number_format = num_fmt_tb()
    ws.cell(row=36, column=1).font = Font(bold=True)

    ws.cell(row=37, column=1, value="Tape purchase (PB)")
    ws.cell(row=37, column=2, value="PB")
    for i in range(N):
        ws.cell(row=37, column=i + 3).value = f"={col(i)}36/1000"
        ws.cell(row=37, column=i + 3).number_format = "#,##0.0"

    # Row 39: TOTAL
    ws.cell(row=39, column=1, value="TOTAL storage purchase (TB)").font = Font(bold=True, size=12)
    ws.cell(row=39, column=2, value="TB")
    for i in range(N):
        c = col(i)
        ws.cell(row=39, column=i + 3).value = f"={c}32+{c}36"
        ws.cell(row=39, column=i + 3).number_format = num_fmt_tb()

    ws.cell(row=40, column=1, value="TOTAL storage purchase (PB)")
    ws.cell(row=40, column=2, value="PB")
    for i in range(N):
        ws.cell(row=40, column=i + 3).value = f"={col(i)}39/1000"
        ws.cell(row=40, column=i + 3).number_format = "#,##0.0"

    # ── LOY1 (FY2026) Purchase Detail ────────────────────────────────
    r = 42
    ws.cell(row=r, column=1, value=f"LOY1 (FY{FY0}) — CURRENT YEAR DETAIL").font = Font(bold=True, size=13, color="C00000")

    ws.cell(row=43, column=1, value="Months elapsed in FY")
    ws.cell(row=43, column=2, value=cur.get("months_elapsed", 3))

    ws.cell(row=44, column=1, value="Months remaining")
    ws.cell(row=44, column=2, value=months_remaining)

    ws.cell(row=45, column=1, value="Currently running")
    ws.cell(row=45, column=2, value="DP2 (10% of DR1)")

    ws.cell(row=47, column=1, value="DP2 compute (core-hours)")
    dp2_ch = f"=ROUND('Ops Compute'!C3*{dp2_frac},0)"
    ws.cell(row=47, column=2).value = dp2_ch
    ws.cell(row=47, column=2).number_format = num_fmt_int()

    ws.cell(row=48, column=1, value="DP2 cores needed")
    drp_py = P["drp"]
    proc_hours_pp = drp_py["processing_window_days"] * 24
    idle_pp = float(drp_py.get("idle_and_inter_stage_multiplier", 1.0))
    dp_dev = P.get("developer_pilot") or {}
    rsp_pp = float(dp_dev.get("usdf_rsp_extra_fraction_of_drp_concurrent", 0))
    drp_core_scale_pp = idle_pp * (1.0 + rsp_pp)
    ws.cell(row=48, column=2).value = (
        f"=ROUND(B47/{proc_hours_pp}*{drp_core_scale_pp:.12g},0)"
    )
    ws.cell(row=48, column=2).number_format = num_fmt_int()

    ws.cell(row=49, column=1, value="Existing batch cores")
    ws.cell(row=49, column=2, value=existing_cores).number_format = num_fmt_int()

    ws.cell(row=50, column=1, value="Compute status for LOY1")
    ws.cell(row=50, column=2).value = '=IF(B49>=B48,"Sufficient — no purchase needed","Gap: "&TEXT(B48-B49,"#,##0")&" cores")'

    ws.cell(row=52, column=1, value="Storage installed (online, PB)")
    ws.cell(row=52, column=2, value=P["current_fleet"]["storage"]["total_installed_pb"])

    ws.cell(row=53, column=1, value="LOY1 online storage needed (TB)")
    ws.cell(row=53, column=2).value = "=C30"
    ws.cell(row=53, column=2).number_format = num_fmt_tb()

    ws.cell(row=54, column=1, value="LOY1 online purchase needed (TB)")
    ws.cell(row=54, column=2).value = "=C32"
    ws.cell(row=54, column=2).number_format = num_fmt_tb()

    ws.cell(row=55, column=1, value="LOY1 online purchase needed (PB)")
    ws.cell(row=55, column=2).value = "=B54/1000"
    ws.cell(row=55, column=2).number_format = "#,##0.1"

    ws.cell(row=56, column=1, value="LOY1 tape purchase needed (TB)")
    ws.cell(row=56, column=2).value = "=C36"
    ws.cell(row=56, column=2).number_format = num_fmt_tb()

    ws.cell(row=58, column=1, value="Purchase summary for FY2026 cycle:").font = Font(bold=True, italic=True)
    ws.cell(row=59, column=1, value="  Batch compute nodes")
    ws.cell(row=59, column=2).value = "=C15"
    ws.cell(row=60, column=1, value="  K8s nodes")
    ws.cell(row=60, column=2).value = "=C17"
    ws.cell(row=61, column=1, value="  Qserv nodes")
    ws.cell(row=61, column=2).value = "=C18"
    ws.cell(row=62, column=1, value="  Online storage (PB)")
    ws.cell(row=62, column=2).value = "=B55"
    ws.cell(row=62, column=2).number_format = "#,##0.1"
    ws.cell(row=63, column=1, value="  Tape storage (PB)")
    ws.cell(row=63, column=2).value = "=B56/1000"
    ws.cell(row=63, column=2).number_format = "#,##0.1"


# ══════════════════════════════════════════════════════════════════════
# International Compute Tab
# ══════════════════════════════════════════════════════════════════════

def build_international_tab(ws, P, N, FY0):
    ws.sheet_properties.tabColor = "00B0F0"
    widths = [40, 14] + [18] * N
    set_col_widths(ws, widths)

    intl = P["international_compute"]
    france = intl["france_drp_share"]
    uk = intl["uk_drp_share"]
    usdf_share = 1.0 - france - uk

    milestones = P.get("milestones_by_loy", {})

    r = 1
    ws.cell(row=r, column=1,
            value="International DRP Compute Sharing").font = Font(bold=True, size=14)
    r = 2
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Unit")
    for i in range(N):
        ws.cell(row=r, column=i + 3, value=f"LOY{i+1}")
    apply_header_style(ws, r, N + 2)

    ws.cell(row=3, column=1, value="Fiscal Year")
    for i in range(N):
        ws.cell(row=3, column=i + 3, value=f"FY{FY0+i}")

    ws.cell(row=4, column=1, value="Milestone")
    for i in range(N):
        ms = milestones.get(i + 1, "")
        ws.cell(row=4, column=i + 3, value=ms)

    # DRP overview
    ws.cell(row=6, column=1, value="DRP COMPUTE SPLIT").font = section_font()

    ws.cell(row=7, column=1, value="Total DRP core-hours")
    ws.cell(row=7, column=2, value="core-hrs")
    for i in range(N):
        ws.cell(row=7, column=i + 3).value = f"='Ops Compute'!{col(i)}3"
        ws.cell(row=7, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=8, column=1, value="Total DRP cores")
    ws.cell(row=8, column=2, value="cores")
    for i in range(N):
        ws.cell(row=8, column=i + 3).value = f"='Ops Compute'!{col(i)}4"
        ws.cell(row=8, column=i + 3).number_format = num_fmt_int()

    shares = [
        (10, f"USDF ({usdf_share*100:.0f}%)", usdf_share),
        (11, f"France / CC-IN2P3 ({france*100:.0f}%)", france),
        (12, f"UK / UKDF ({uk*100:.0f}%)", uk),
    ]

    ws.cell(row=9, column=1, value="DRP Cores by Facility").font = Font(bold=True)
    for row_n, label, share in shares:
        ws.cell(row=row_n, column=1, value=label)
        ws.cell(row=row_n, column=2, value="cores")
        for i in range(N):
            ws.cell(row=row_n, column=i + 3).value = f"=ROUND({col(i)}8*{share},0)"
            ws.cell(row=row_n, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=14, column=1, value="DRP Core-Hours by Facility").font = Font(bold=True)
    for idx, (_, label, share) in enumerate(shares):
        row_n = 15 + idx
        ws.cell(row=row_n, column=1, value=label)
        ws.cell(row=row_n, column=2, value="core-hrs")
        for i in range(N):
            ws.cell(row=row_n, column=i + 3).value = f"=ROUND({col(i)}7*{share},0)"
            ws.cell(row=row_n, column=i + 3).number_format = num_fmt_int()

    # France detail
    ws.cell(row=19, column=1, value="FRANCE (CC-IN2P3) DETAIL").font = section_font()

    eff = eff_torino_cores(P)
    ws.cell(row=20, column=1, value="DRP cores")
    ws.cell(row=20, column=2, value="cores")
    for i in range(N):
        ws.cell(row=20, column=i + 3).value = f"={col(i)}11"
        ws.cell(row=20, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=21, column=1, value="Batch nodes (est.)")
    ws.cell(row=21, column=2, value="nodes")
    for i in range(N):
        ws.cell(row=21, column=i + 3).value = f"=CEILING({col(i)}20/{eff},1)"
        ws.cell(row=21, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=22, column=1, value="DRP scratch storage (est.)")
    ws.cell(row=22, column=2, value="TB")
    for i in range(N):
        ws.cell(row=22, column=i + 3).value = (
            f"=ROUND(('Ops Storage'!{col(i)}38+'Ops Storage'!{col(i)}39"
            f"+'Ops Storage'!{col(i)}41)*{france},0)"
        )
        ws.cell(row=22, column=i + 3).number_format = num_fmt_tb()

    # UK detail
    ws.cell(row=24, column=1, value="UK (UKDF) DETAIL").font = section_font()

    ws.cell(row=25, column=1, value="DRP cores")
    ws.cell(row=25, column=2, value="cores")
    for i in range(N):
        ws.cell(row=25, column=i + 3).value = f"={col(i)}12"
        ws.cell(row=25, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=26, column=1, value="Batch nodes (est.)")
    ws.cell(row=26, column=2, value="nodes")
    for i in range(N):
        ws.cell(row=26, column=i + 3).value = f"=CEILING({col(i)}25/{eff},1)"
        ws.cell(row=26, column=i + 3).number_format = num_fmt_int()

    ws.cell(row=27, column=1, value="DRP scratch storage (est.)")
    ws.cell(row=27, column=2, value="TB")
    for i in range(N):
        ws.cell(row=27, column=i + 3).value = (
            f"=ROUND(('Ops Storage'!{col(i)}38+'Ops Storage'!{col(i)}39"
            f"+'Ops Storage'!{col(i)}41)*{uk},0)"
        )
        ws.cell(row=27, column=i + 3).number_format = num_fmt_tb()

    ws.cell(row=29, column=1,
            value="Note: France and UK handle DRP compute only. Long-term data products "
                  "and storage reside at USDF.").font = Font(italic=True, color="666666")


# ══════════════════════════════════════════════════════════════════════
# Yearly Readiness Tab (expanded per-LOY readiness)
# ══════════════════════════════════════════════════════════════════════

def build_yearly_readiness_tab(ws, P, N, FY0):
    ws.sheet_properties.tabColor = "7030A0"
    set_col_widths(ws, [18, 14, 12, 20, 20, 20, 18, 18, 18, 38])

    intl = P["international_compute"]
    france = intl["france_drp_share"]
    uk = intl["uk_drp_share"]

    milestones = P.get("milestones_by_loy", {})
    installed_pb = P["current_fleet"]["storage"]["total_installed_pb"]
    fleet = P["current_fleet"]["batch"]
    existing_cores = (fleet["milano_nodes"] * fleet["milano_cores_per_node"] +
                      fleet["torino_nodes"] * fleet["torino_cores_per_node"])

    r = 1
    ws.cell(row=r, column=1,
            value="Yearly Readiness — LOY1 to LOY10").font = Font(bold=True, size=14)

    r = 3
    headers = ["LOY", "Year", "Milestone",
               "Total Cores\nNeeded",
               "Cores Available\n(before purchase)",
               "Compute\nGap",
               "Online Storage\nNeeded (PB)",
               "Storage Available\n(before purchase, PB)",
               "Storage\nGap (PB)",
               "Status"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci, value=h)
    apply_header_style(ws, r, len(headers))

    for i in range(N):
        r = 4 + i
        loy = i + 1
        c = col(i)
        ms = milestones.get(loy, "")
        fy = FY0 + i

        ws.cell(row=r, column=1, value=f"LOY{loy}")
        ws.cell(row=r, column=2, value=fy)
        ms_cell = ws.cell(row=r, column=3, value=ms)
        if ms:
            ms_cell.font = Font(bold=True, color="C00000")

        # Total USDF cores needed (100% — same as Purchase Plan row 11)
        ws.cell(row=r, column=4).value = f"='Purchase Plan'!{c}11"
        ws.cell(row=r, column=4).number_format = num_fmt_int()

        # Compute available BEFORE this year's purchase
        if i == 0:
            ws.cell(row=r, column=5, value=existing_cores)
        else:
            ws.cell(row=r, column=5).value = f"='Purchase Plan'!{col(i-1)}12"
        ws.cell(row=r, column=5).number_format = num_fmt_int()

        # Compute gap (positive = shortfall)
        ws.cell(row=r, column=6).value = f"=MAX(0,D{r}-E{r})"
        ws.cell(row=r, column=6).number_format = num_fmt_int()

        # Storage needed on floor (PB)
        ws.cell(row=r, column=7).value = f"='Purchase Plan'!{c}30/1000"
        ws.cell(row=r, column=7).number_format = "#,##0.0"

        # Storage available BEFORE this year's purchase (PB)
        if i == 0:
            ws.cell(row=r, column=8, value=installed_pb)
        else:
            ws.cell(row=r, column=8).value = f"='Purchase Plan'!{col(i-1)}31/1000"
        ws.cell(row=r, column=8).number_format = "#,##0.0"

        # Storage gap (positive = shortfall)
        ws.cell(row=r, column=9).value = f"=MAX(0,G{r}-H{r})"
        ws.cell(row=r, column=9).number_format = "#,##0.0"

        # Status
        ws.cell(row=r, column=10).value = (
            f'=IF(AND(F{r}<=0,I{r}<=0),"On Track",'
            f'"Need: "&IF(F{r}>0,TEXT(F{r},"#,##0")&" cores","")&'
            f'IF(AND(F{r}>0,I{r}>0)," + ","")&'
            f'IF(I{r}>0,TEXT(I{r},"#,##0.0")&" PB",""))'
        )

    r = 4 + N + 1
    ws.cell(row=r, column=1, value="Notes:").font = Font(bold=True, italic=True)
    ws.cell(row=r + 1, column=1,
            value="'Available' = what is on hand BEFORE this year's purchase cycle. "
                  "'Gap' = additional capacity to procure. 100% USDF compute assumed."
            ).font = Font(italic=True, color="666666")
    ws.cell(row=r + 2, column=1,
            value=f"See 'International Compute' tab for France ({france*100:.0f}%) + UK ({uk*100:.0f}%) DRP offload scenario."
            ).font = Font(italic=True, color="666666")


# ══════════════════════════════════════════════════════════════════════
# Named Ranges
# ══════════════════════════════════════════════════════════════════════

def define_named_ranges(wb, P):
    from openpyxl.workbook.defined_name import DefinedName

    ranges = {
        "cpuFac": "'Model'!$C$19",
        "diskFac": "'Model'!$C$20",
        "qservFac": "'Model'!$C$21",
        "scaleYear": "'Model'!$B$22",
        "computeLifetime": "'Model'!$B$23",
        "storageLifetime": "'Model'!$B$24",
    }

    for name, ref in ranges.items():
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names.add(dn)


# ══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    main()
