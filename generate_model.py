#!/usr/bin/env python3
"""
Build the data-facility sizing and pricing workbook.

    uv run generate_model.py

Reads every number, label and note from ``sizing_params.yaml`` and writes an
Excel workbook of live formulas covering the loop years of operations.

DESIGN
    This script holds formulas and layout mechanics only. It contains no
    facility data: everything a reader sees — tab titles, section banners,
    row labels, units, notes — is declared in the ``workbook:`` block of the
    parameter file, and every value is either a literal there or a dotted
    pointer into the parameter blocks above it.

    Row numbers appear nowhere. They are derived from the order of the YAML:
    a section banner, then its rows, then one blank line. Formulas address
    other rows by key, so inserting or removing a row re-points every
    dependent formula automatically.

    A facility adopting this model should need to edit only the YAML.

TAB ARCHITECTURE (enforced; no tab duplicates another tab's numbers)

    Key Numbers      REFERENCE. Sizing and data-product inputs.
    Cost Inputs      REFERENCE. Costs only.
    Projection tab   CALCULATED from Key Numbers.
    Facility Split   CALCULATED from Key Numbers + the projection tab.
    Qserv            CALCULATED from Key Numbers.
    DF Template      CALCULATED from Key Numbers + the projection tab.
    Pricing Forecast CALCULATED from Key Numbers + Cost Inputs, reusing demand
                     from the projection tab and node counts from Qserv.

CELL COLOUR CONVENTION
    blue    raw input, safe to edit
    black   formula
    green   link to another tab
    yellow  key assumption that materially moves the answer
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ─────────────────────────── styling ────────────────────────────
FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")
BLACK = Font(name=FONT, size=10)
GREEN = Font(name=FONT, size=10, color="008000")
GRAY_IT = Font(name=FONT, size=9, italic=True, color="808080")
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=13, bold=True)
SECTION = Font(name=FONT, size=10, bold=True, color="FFFFFF")
ITALIC = Font(name=FONT, size=9, italic=True, color="666666")
YELLOW = PatternFill("solid", fgColor="FFFF00")
SECTION_FILL = PatternFill("solid", fgColor="4472C4")
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))

FORMATS = {
    "int": "#,##0", "dec1": "#,##0.0", "dec2": "#,##0.00", "dec3": "#,##0.000",
    "pct": "0%", "pct2": "0.00%", "usd": "$#,##0", "usdm": '$#,##0.00,,"M"',
    "tb8": "0.00000000", "ratio3": "0.000",
}
FONT_FOR_STYLE = {"link": GREEN, "sub": GRAY_IT, "bold": BLACK, None: BLACK}

DEFAULT_PARAMS = "sizing_params.yaml"

# Arithmetic, not facility data.
HOURS_PER_DAY = 24
TB_PER_PB = 1000
BYTES_PER_TB = 1_000_000_000_000
PER_BILLION = 1_000_000_000
PER_MILLION = 1_000_000


# ──────────────────────── small helpers ─────────────────────────

def dotted(params: dict, path: str):
    """Resolve 'a.b.c' against the parameter file."""
    node = params
    for part in path.split("."):
        if not isinstance(node, dict) or part not in node:
            raise KeyError(f"parameter path not found: {path!r} (at {part!r})")
        node = node[part]
    return node


def col(i: int) -> str:
    """Loop year i (0-based) occupies column C onwards on projection tabs."""
    return chr(ord("C") + i)


def prev(c: str) -> str:
    return chr(ord(c) - 1)


def idx(c: str) -> int:
    return ord(c) - ord("C")


def check_note(note) -> str:
    """A note beginning with '=' would be parsed as a formula by Excel."""
    text = "" if note is None else str(note)
    if text.startswith("="):
        raise ValueError(f"note must not start with '=': {text[:60]!r}")
    return text


def banner(ws, row: int, text: str, ncols: int) -> None:
    ws.cell(row=row, column=1, value=text).font = SECTION
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL


def note_block(ws, row: int, text: str, ncols: int) -> None:
    cell = ws.cell(row=row, column=1, value=check_note(text))
    cell.font = ITALIC
    cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 26 if len(str(text)) > 95 else 14


# ────────────────────────── layout pass ─────────────────────────
# Row numbers are derived once, for every tab, before anything is written.
# Rule: a section banner, then any header line the section needs, then its
# rows, then one blank line.

def lay_out(spec: dict, start_row: int, params: dict) -> tuple[dict, list]:
    row = start_row
    rowmap: dict[str, int] = {}
    plan: list[tuple] = []
    for sec in spec.get("sections", []):
        kind = sec.get("kind", "rows")
        plan.append(("section", row, sec))
        row += 1
        if kind in ("history_grid", "dataset_table"):
            plan.append((f"{kind}_header", row, sec))
            row += 1
        for r in sec.get("rows", []):
            rowmap[r["key"]] = row
            plan.append(("row", row, r))
            row += 1
        if kind == "dataset_table":
            entries = dotted(params, sec["source"])
            plan.append(("dataset_body", row, sec))
            row += len(entries)
            rowmap["ds_other"] = row
            plan.append(("dataset_other", row, sec))
            row += 1
            rowmap[sec["total_key"]] = row
            plan.append(("dataset_total", row, sec))
            row += 1
        row += 1
    return rowmap, plan


def last_row_of(plan: list) -> int:
    return max(row for _, row, _ in plan)


# ─────────────────────── reference tabs ─────────────────────────

def render_reference(wb, spec, params, widths, rowmap, plan, formulas, ncols,
                     note_col):
    """Key Numbers / Cost Inputs: label, value, unit, note."""
    ws = wb.create_sheet(spec["tab"])
    if spec.get("tab_color"):
        ws.sheet_properties.tabColor = spec["tab_color"]
    for width, letter in zip(widths, "ABCDEFG"):
        ws.column_dimensions[letter].width = width
    ws["A1"] = spec["title"]
    ws["A1"].font = TITLE
    ws["A2"] = spec["intro"]
    ws["A2"].font = ITALIC

    for kind, row, payload in plan:
        if kind == "section":
            banner(ws, row, payload["title"], ncols)

        elif kind == "history_grid_header":
            history = dotted(params, payload["source"])
            first_fy = history["first_fiscal_year"]
            width = len(next(v for v in history.values() if isinstance(v, list)))
            for i in range(width):
                cell = ws.cell(row=row, column=2 + i, value=f"FY{first_fy + i}")
                cell.font = BOLD
                cell.fill = HDR_FILL

        elif kind == "dataset_table_header":
            for c, label in enumerate(payload["header"], start=1):
                cell = ws.cell(row=row, column=c, value=label)
                cell.font = BOLD
                cell.fill = HDR_FILL
            ws.cell(row=row, column=4,
                    value=check_note(payload["header_note"])).font = ITALIC

        elif kind == "dataset_body":
            for i, (name, size, count) in enumerate(dotted(params, payload["source"])):
                ws.cell(row=row + i, column=1, value=name).font = BLACK
                for c, value, fmt in ((2, size, "dec3"), (3, count, "int")):
                    cell = ws.cell(row=row + i, column=c, value=value)
                    cell.number_format = FORMATS[fmt]
                    cell.font = BLUE

        elif kind == "dataset_other":
            ws.cell(row=row, column=1, value=payload["other_label"]).font = BLACK
            for c, fmt in ((2, "dec3"), (3, "int")):
                cell = ws.cell(row=row, column=c, value=0)
                cell.number_format = FORMATS[fmt]
                cell.font = BLUE
            ws.cell(row=row, column=4,
                    value=check_note(payload["other_note"])).font = ITALIC

        elif kind == "dataset_total":
            body_start = rowmap["ds_other"] - len(dotted(params, payload["source"]))
            ws.cell(row=row, column=1, value=payload["total_label"]).font = BOLD
            for c, fmt in ((2, "dec1"), (3, "int")):
                letter = chr(64 + c)
                cell = ws.cell(row=row, column=c,
                               value=f"=SUM({letter}{body_start}:"
                                     f"{letter}{rowmap['ds_other']})")
                cell.number_format = FORMATS[fmt]
                cell.font = BLACK
            ws.cell(row=row, column=4,
                    value=check_note(payload["total_note"])).font = ITALIC
            note_block(ws, row + 2, payload["note"], ncols)

        else:  # ordinary row
            r = payload
            ws.cell(row=row, column=1, value=r["label"]).font = BLACK

            if "series" in r:  # historic-purchase grid line
                for i, value in enumerate(dotted(params, r["_source"] + "." + r["series"])):
                    cell = ws.cell(row=row, column=2 + i, value=value)
                    cell.number_format = FORMATS["int"]
                    cell.font = BLUE
                    cell.fill = YELLOW
                ws.cell(row=row, column=note_col,
                        value=check_note(r.get("note"))).font = ITALIC
                continue

            if r.get("formula"):
                value, is_formula = formulas[r["key"]](), True
            elif "param" in r:
                value = dotted(params, r["param"])
                if "scale" in r:
                    value *= r["scale"]
                is_formula = False
            else:
                value, is_formula = r.get("value"), False

            cell = ws.cell(row=row, column=2, value=value)
            cell.font = BLACK if is_formula else BLUE
            if r.get("format"):
                cell.number_format = FORMATS[r["format"]]
            if r.get("yellow"):
                cell.fill = YELLOW
            ws.cell(row=row, column=3, value=r.get("unit")).font = BLACK
            ws.cell(row=row, column=note_col,
                    value=check_note(r.get("note"))).font = ITALIC
            for c in range(1, note_col + 1):
                ws.cell(row=row, column=c).border = THIN

    for i, text in enumerate(spec.get("notes", [])):
        note_block(ws, last_row_of(plan) + 2 + i, text, ncols)
    return ws


# ─────────────────────── projection tabs ────────────────────────

def projection_header(ws, spec, n_loy, fy0, milestones, header_row):
    """LOY / fiscal-year / milestone header. The first loop year is annotated
    as an actual unless the tab opts out with mark_first_actual: false."""
    mark_actual = spec.get("mark_first_actual", True)
    ws.cell(row=header_row, column=1, value="Metric").font = BOLD
    ws.cell(row=header_row, column=2, value="Unit").font = BOLD
    for i in range(n_loy):
        c = 3 + i
        ws.cell(row=header_row, column=c, value=f"LOY{i + 1}").font = BOLD
        ws.cell(row=header_row + 1, column=c, value=f"FY{fy0 + i}").font = BLACK
        milestone = milestones.get(i + 1, "")
        if i == 0 and mark_actual:
            milestone = f"{milestone} (actual)" if milestone else "(actual)"
        ws.cell(row=header_row + 2, column=c, value=milestone).font = ITALIC
    ws.cell(row=header_row + 1, column=1, value="Fiscal Year").font = BLACK
    ws.cell(row=header_row + 2, column=1, value="Milestone").font = BLACK
    for c in range(1, 3 + n_loy):
        ws.cell(row=header_row, column=c).fill = HDR_FILL


def new_projection_sheet(wb, spec, n_loy):
    ws = wb.create_sheet(spec["tab"])
    if spec.get("tab_color"):
        ws.sheet_properties.tabColor = spec["tab_color"]
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 12
    for i in range(n_loy):
        ws.column_dimensions[col(i)].width = 13
    ws["A1"] = spec["title"]
    ws["A1"].font = TITLE
    return ws


def render_projection(ws, spec, params, plan, n_loy, formulas, firsts,
                      total_col=None):
    """Write one row per YAML entry across every loop year."""
    ncols = total_col or n_loy + 2
    for kind, row, payload in plan:
        if kind == "section":
            banner(ws, row, payload["title"], ncols)
            continue
        r = payload
        style = r.get("style")
        value_font = FONT_FOR_STYLE.get(style, BLACK)
        label_font = GRAY_IT if style == "sub" else (BOLD if style == "bold" else BLACK)
        fmt = FORMATS[r.get("format", "int")]
        row_kind = r.get("kind", "grid")

        if row_kind == "note":
            ws.cell(row=row, column=1, value=check_note(r["label"])).font = ITALIC
            continue

        ws.cell(row=row, column=1, value=r["label"]).font = label_font
        ws.cell(row=row, column=2, value=r.get("unit")).font = label_font

        if row_kind == "series":
            values = dotted(params, r["param"])
            divisor = r.get("divisor", 1)
            editable = r.get("yellow", r["key"].startswith("sh_"))
            for i in range(n_loy):
                cell = ws.cell(row=row, column=3 + i, value=values[i] / divisor)
                cell.number_format = fmt
                cell.font = BLUE
                if editable:
                    cell.fill = YELLOW
            continue

        if row_kind == "scalar":
            ws.cell(row=row, column=1).font = BOLD
            cell = ws.cell(row=row, column=2, value=formulas[r["key"]]("B"))
            cell.number_format = fmt
            if r.get("note"):
                ws.cell(row=row, column=4, value=check_note(r["note"])).font = ITALIC
            continue

        for i in range(n_loy):
            c = col(i)
            value = firsts[r["key"]] if (i == 0 and r["key"] in firsts) \
                else formulas[r["key"]](c)
            if value == "":
                continue
            cell = ws.cell(row=row, column=3 + i, value=value)
            cell.number_format = fmt
            cell.font = value_font
        if r.get("total") and total_col:
            cell = ws.cell(row=row, column=total_col,
                           value=f"=SUM(C{row}:{col(n_loy - 1)}{row})")
            cell.number_format = fmt
            cell.font = BOLD

    for i, text in enumerate(spec.get("notes", [])):
        note_block(ws, last_row_of(plan) + 2 + i, text, ncols)


# ───────────────────────── the model ────────────────────────────

def build_workbook(P: dict, output_path: str | None = None) -> str:
    W = P["workbook"]
    output_path = output_path or W["output"]
    n_loy = P["general"]["num_loy"]
    fy0 = P["general"]["start_fiscal_year"]
    milestones = P["milestones_by_loy"]
    last_col = col(n_loy - 1)

    KN, CI = f"'{W['key_numbers']['tab']}'", f"'{W['cost_inputs']['tab']}'"
    AU, QS = f"'{W['all_at_usdf']['tab']}'", f"'{W['qserv']['tab']}'"

    # Tag history-grid rows with their source so the renderer can find them.
    for sec in W["cost_inputs"]["sections"]:
        if sec.get("kind") == "history_grid":
            for r in sec["rows"]:
                r["_source"] = sec["source"]

    # ── layout pass: every row number, before anything is written ──
    qs_doc = P["qserv_documentation"]["entries"]
    qs_header = 3 + len(qs_doc) + 2
    qs_start = qs_header + 4

    kn_map, kn_plan = lay_out(W["key_numbers"], 3, P)
    ci_map, ci_plan = lay_out(W["cost_inputs"], 3, P)
    au_map, au_plan = lay_out(W["all_at_usdf"], 6, P)
    fs_map, fs_plan = lay_out(W["facility_split"], 6, P)
    qs_map, qs_plan = lay_out(W["qserv"], qs_start, P)
    pf_map, pf_plan = lay_out(W["pricing_forecast"], 5, P)

    def K(key):
        return f"{KN}!$B${kn_map[key]}"

    def C(key):
        return f"{CI}!$B${ci_map[key]}"

    def H(key):
        return f"{CI}!$B${ci_map[key]}:$F${ci_map[key]}"

    A, F, Q, R = au_map.__getitem__, fs_map.__getitem__, qs_map.__getitem__, pf_map.__getitem__

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Key Numbers ─────────────────────────────────────────────
    def B(key):
        return f"B{kn_map[key]}"

    kn_formulas = {
        "input_tb": lambda: f"={B('nights')}*{B('visits')}*{B('img_tb')}*{B('raw_comp')}",
        "bytes_written": lambda: f"={B('ds_total')}",
        "intermediates": lambda: f"={B('bytes_written')}-{B('retained')}",
        "batch_nodes": lambda: f"={B('milano')}+{B('torino')}",
        "batch_cores": lambda: f"={B('batch_nodes')}*{B('cores_node')}",
        "milano_eq": lambda: (f"=({B('milano')}+{B('torino')}*{B('speedup')})"
                              f"*{B('cores_node')}"),
        "usdf_share": lambda: f"=1-{B('fr_share')}-{B('uk_share')}",
    }
    render_reference(wb, W["key_numbers"], P, [50, 16, 14, 46],
                     kn_map, kn_plan, kn_formulas, ncols=4, note_col=4)

    # ── Projection tab ──────────────────────────────────────────
    au_formulas = {
        "survey_years": lambda c: idx(c),
        "input_tb": lambda c: f"={c}{A('survey_years')}*{K('input_tb')}",
        "core_hours": lambda c: (f"={c}{A('input_tb')}*{K('ch_per_tb')}"
                                 f"+{c}{A('survey_years')}*{K('pilot_nd')}"
                                 f"*{K('cores_node')}*{HOURS_PER_DAY}"),
        "node_days": lambda c: f"={c}{A('core_hours')}/({HOURS_PER_DAY}*{K('cores_node')})",
        "peak_cores": lambda c: (f"=ROUND({c}{A('core_hours')}"
                                 f"/({K('window')}*{HOURS_PER_DAY})"
                                 f"*{K('idle')}*(1+{K('rsp')}),0)"),
        "ch_safety": lambda c: f"={c}{A('core_hours')}*{K('safety')}",
        "nodes_needed": lambda c: f"=ROUNDUP({c}{A('peak_cores')}/{K('cores_node')},0)",
        "fleet_eq": lambda c: f"={K('milano_eq')}",
        "utilization": lambda c: f"={c}{A('peak_cores')}/{c}{A('fleet_eq')}",
        "written_pb": lambda c: f"={K('bytes_written')}*{c}{A('survey_years')}",
        "written_tb": lambda c: f"={c}{A('written_pb')}*{TB_PER_PB}",
        "rel_pb": lambda c: f"={K('retained')}*{c}{A('survey_years')}",
        "inter_pb": lambda c: f"={c}{A('written_pb')}-{c}{A('rel_pb')}",
        "live_inter": lambda c: f"={c}{A('inter_pb')}*{K('livef')}",
        "on_rel": lambda c: (f"={K('retained')}*({c}{A('survey_years')}"
                             f"*MIN({c}{A('survey_years')},{K('retwin')})"
                             f"-(MIN({c}{A('survey_years')},{K('retwin')})-1)"
                             f"*MIN({c}{A('survey_years')},{K('retwin')})/2)"),
        "on_raw": lambda c: f"={K('rawcal')}*(COLUMN()-2)/{TB_PER_PB}",
        "on_prompt": lambda c: f"=({K('t1')}*(COLUMN()-2)+{K('t2')})/{TB_PER_PB}",
        "subtotal": lambda c: f"=SUM({c}{A('live_inter')}:{c}{A('on_prompt')})",
        "onfloor_pb": lambda c: f"={c}{A('subtotal')}*(1+{K('other')})",
        "onfloor_tb": lambda c: f"={c}{A('onfloor_pb')}*{TB_PER_PB}",
        "installed": lambda c: f"={K('installed_pb')}",
        "tape_raw": lambda c: (f"={K('rawcal')}*{K('tape_copies')}*(COLUMN()-2)"
                               f"/{TB_PER_PB}"),
        "tape_relc": lambda c: (f"={K('retained')}*{c}{A('survey_years')}"
                                f"*({c}{A('survey_years')}+1)/2*{K('tape_rel')}"),
        "tape_bk": lambda c: f"={c}{A('onfloor_pb')}*{K('tape_backup')}",
        "tape_fp": lambda c: (f"=SUM({c}{A('tape_raw')}:{c}{A('tape_bk')})"
                              f"+{K('tape_pre')}/{TB_PER_PB}"),
        "tape_fp_tb": lambda c: f"={c}{A('tape_fp')}*{TB_PER_PB}",
        "tape_avail": lambda c: f"={K('tape_owned')}",
        "tape_buy_cum": lambda c: f"=MAX(0,{c}{A('tape_fp_tb')}-{c}{A('tape_avail')})",
        "tape_buy": lambda c: f"={c}{A('tape_buy_cum')}-{prev(c)}{A('tape_buy_cum')}",
        "ap": lambda c: f"={K('ap_cores')}",
        "ap_nodes": lambda c: f"=ROUNDUP({c}{A('ap')}/{K('k8s_cores_node')},0)",
        "dac": lambda c: f"=ROUND({c}{A('peak_cores')}*{K('dac_frac')},0)",
        "staff": lambda c: f"=ROUND({c}{A('dac')}*{K('staff_frac')},0)",
        "interactive": lambda c: f"={K('int_nodes')}",
        "svc_util": lambda c: (f"=({c}{A('ap')}+{c}{A('dac')}+{c}{A('staff')})"
                               f"/({K('k8s_nodes')}*{K('k8s_cores_node')})"),
        "apdb_tb": lambda c: (f"={K('apdb_ref_tb')}/{K('apdb_ref_visits')}"
                              f"*{K('nights')}*{K('visits')}*(COLUMN()-2)"),
        "apdb_bk": lambda c: f"={K('apdb_backup')}/{n_loy}*(COLUMN()-2)",
        "cass": lambda c: f"={K('cass_nodes')}",
        "files_written": lambda c: (f"={KN}!$C${kn_map['ds_total']}/{PER_MILLION}"
                                    f"*{c}{A('survey_years')}"),
        "files_floor": lambda c: (f"={c}{A('files_written')}*({K('livef')}"
                                  f"*(1-{K('retained')}/{K('bytes_written')})"
                                  f"+{K('retained')}/{K('bytes_written')})"),
    }
    au_firsts = {
        "survey_years": f"={K('dp2_visits')}/({K('nights')}*{K('visits')})",
        "core_hours": f"={K('dp2_nd')}*{K('cores_node')}*{HOURS_PER_DAY}",
        "written_pb": f"={K('dp2_out')}",
        "rel_pb": f"={K('dp2_rel')}",
        "tape_relc": f"={K('dp2_rel')}*{K('tape_rel')}",
        "tape_buy": f"=C{A('tape_buy_cum')}",
    }
    ws_au = new_projection_sheet(wb, W["all_at_usdf"], n_loy)
    projection_header(ws_au, W["all_at_usdf"], n_loy, fy0, milestones, 2)
    render_projection(ws_au, W["all_at_usdf"], P, au_plan, n_loy,
                      au_formulas, au_firsts)
    ws_au.cell(row=A("survey_years"), column=3).font = GREEN  # it is a link

    # ── Facility Split ──────────────────────────────────────────
    fs_formulas = {
        "sh_local": lambda c: f"=1-{c}{F('sh_fr')}-{c}{F('sh_uk')}",
        "uk_agg": lambda c: (f"=SUMPRODUCT(D{F('sh_uk')}:{last_col}{F('sh_uk')},"
                             f"{AU}!D{A('survey_years')}:{last_col}{A('survey_years')})"
                             f"/SUM({AU}!D{A('survey_years')}:"
                             f"{last_col}{A('survey_years')})"),
        "tot_ch": lambda c: f"={AU}!{c}{A('core_hours')}",
        "ch_local": lambda c: f"={c}{F('tot_ch')}*{c}{F('sh_local')}",
        "ch_fr": lambda c: f"={c}{F('tot_ch')}*{c}{F('sh_fr')}",
        "ch_uk": lambda c: f"={c}{F('tot_ch')}*{c}{F('sh_uk')}",
        "nd_local": lambda c: f"={c}{F('ch_local')}/({HOURS_PER_DAY}*{K('cores_node')})",
        "nd_fr": lambda c: f"={c}{F('ch_fr')}/({HOURS_PER_DAY}*{K('cores_node')})",
        "nd_uk": lambda c: f"={c}{F('ch_uk')}/({HOURS_PER_DAY}*{K('cores_node')})",
        "pc_local": lambda c: f"=ROUND({AU}!{c}{A('peak_cores')}*{c}{F('sh_local')},0)",
        "pc_fr": lambda c: f"=ROUND({AU}!{c}{A('peak_cores')}*{c}{F('sh_fr')},0)",
        "pc_uk": lambda c: f"=ROUND({AU}!{c}{A('peak_cores')}*{c}{F('sh_uk')},0)",
        "chk_local": lambda c: f"={c}{F('pc_local')}/{K('milano_eq')}",
        "chk_uk": lambda c: f"={c}{F('ch_uk')}/{K('uk_cap')}",
        "st_fr": lambda c: f"={AU}!{c}{A('live_inter')}*{c}{F('sh_fr')}",
        "st_uk": lambda c: f"={AU}!{c}{A('live_inter')}*{c}{F('sh_uk')}",
        "st_arch": lambda c: f"={AU}!{c}{A('on_rel')}",
    }
    ws_fs = new_projection_sheet(wb, W["facility_split"], n_loy)
    projection_header(ws_fs, W["facility_split"], n_loy, fy0, milestones, 2)
    render_projection(ws_fs, W["facility_split"], P, fs_plan, n_loy, fs_formulas, {})

    # ── Qserv (documentation block, then the projection) ────────
    ws_qs = new_projection_sheet(wb, W["qserv"], n_loy)
    banner(ws_qs, 3,
           f"{W['qserv']['spec_title']} (as of {P['qserv_documentation']['as_of']})"
           " — documentation, not calculation inputs", n_loy + 2)
    for i, (label, text) in enumerate(qs_doc):
        row = 4 + i
        ws_qs.cell(row=row, column=1, value=label).font = BOLD
        cell = ws_qs.cell(row=row, column=3, value=text)
        cell.font = BLACK
        cell.alignment = Alignment(wrap_text=True)
        ws_qs.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        ws_qs.row_dimensions[row].height = 28
    projection_header(ws_qs, W["qserv"], n_loy, fy0, milestones, qs_header)

    win = P["storage_retention"]["qserv_window_years"]
    qs_formulas = {
        "czar": lambda c: (f"={c}{Q('objects')}*{PER_BILLION}*{K('obj_row')}"
                           f"*{K('qs_rep')}/{BYTES_PER_TB}"),
        "worker": lambda c: (f"=({c}{Q('objects')}*({K('obj_row')}+{K('obj_extra')})"
                             f"+{c}{Q('sources')}*{K('src_row')}"
                             f"+{c}{Q('fsources')}*{K('fsrc_row')})"
                             f"*{PER_BILLION}*{K('qs_rep')}/{BYTES_PER_TB}"),
        "worker_win": lambda c: "=" + "+".join(
            f"{col(j)}{Q('worker')}"
            for j in range(max(0, idx(c) - (win - 1)), idx(c) + 1)),
        "worker_pb": lambda c: f"={c}{Q('worker_win')}/{TB_PER_PB}",
        "nodes_req": lambda c: f"=ROUNDUP({c}{Q('worker_win')}/{c}{Q('per_node')},0)",
        "nodes_have": lambda c: f"={K('qs_nodes')}",
        "surplus": lambda c: f"={c}{Q('nodes_have')}-{c}{Q('nodes_req')}",
    }
    render_projection(ws_qs, W["qserv"], P, qs_plan, n_loy, qs_formulas, {})

    # ── DF Template ─────────────────────────────────────────────
    render_df_template(wb, W["df_template"], P, K, AU, A, n_loy, fy0, last_col)

    # ── Cost Inputs ─────────────────────────────────────────────
    render_reference(wb, W["cost_inputs"], P, [44, 14, 12, 12, 12, 12, 44],
                     ci_map, ci_plan, {}, ncols=7,
                     note_col=W["cost_inputs"]["note_col"])

    # ── Pricing Forecast ────────────────────────────────────────
    render_pricing(wb, W["pricing_forecast"], P, pf_plan, n_loy, fy0, milestones,
                   K, C, H, R, A, Q, AU, QS, last_col)

    wb.save(output_path)
    return output_path


def render_df_template(wb, spec, P, K, AU, A, n_loy, fy0, last_col):
    ex = P["df_template_example"]
    ws = wb.create_sheet(spec["tab"])
    ws.sheet_properties.tabColor = spec["tab_color"]
    for width, letter in zip([48, 14, 16, 14, 14], "ABCDE"):
        ws.column_dimensions[letter].width = width
    ws["A1"] = spec["title"]
    ws["A1"].font = TITLE
    ws["A3"] = spec["loop_year_label"]
    ws["A3"].font = BOLD
    ws["B3"] = ex["default_loop_year"]
    ws["B3"].font = BLUE
    ws["B3"].fill = YELLOW
    loy = ex["default_loop_year"]
    milestone = P["milestones_by_loy"].get(loy, "")
    ws["C3"] = f"LOY{loy} = FY{fy0 + loy - 1}" + (f" = {milestone}" if milestone else "")
    ws["C3"].font = ITALIC

    for c, label in enumerate(spec["columns"], start=1):
        cell = ws.cell(row=5, column=c, value=label)
        cell.font = BOLD
        cell.fill = HDR_FILL

    peer = ex["peer_cores_per_node"]
    column_values = {
        "share": [f"={K('usdf_share')}", f"={K('fr_share')}", f"={K('uk_share')}",
                  ex["share"]],
        "cores": [f"={K('cores_node')}", peer, peer, ex["cores_per_node"]],
        "window": [f"={K('window')}"] * 3 + [ex["window_days"]],
        "capacity": [f"={K('milano_eq')}*{K('window')}*{HOURS_PER_DAY}", "",
                     f"={K('uk_cap')}", ex["provisioned_ch_per_year"]],
    }
    rowmap, row = {}, 6
    for r in spec["inputs"]:
        rowmap[r["key"]] = row
        ws.cell(row=row, column=1, value=r["label"]).font = BLACK
        for c, value in enumerate(column_values[r["key"]], start=2):
            cell = ws.cell(row=row, column=c, value=value)
            cell.number_format = FORMATS[r["format"]]
            if isinstance(value, str) and value.startswith("="):
                cell.font = GREEN
            else:
                cell.font = BLUE
                cell.fill = YELLOW
        row += 1

    ws.cell(row=row, column=1, value=spec["computed_title"]).font = SECTION
    for c in range(1, len(spec["columns"]) + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    row += 1
    for r in spec["computed"]:
        rowmap[r["key"]] = row
        row += 1
    D = rowmap.__getitem__

    formulas = {
        "total_ch": lambda c: (f"=INDEX({AU}!$C${A('core_hours')}:"
                               f"${last_col}${A('core_hours')},1,$B$3)"),
        "your_ch": lambda c: f"={c}{D('total_ch')}*{c}{D('share')}",
        "your_nd": lambda c: f"={c}{D('your_ch')}/({HOURS_PER_DAY}*{c}{D('cores')})",
        "your_pc": lambda c: (f"=ROUND({c}{D('your_ch')}"
                              f"/({c}{D('window')}*{HOURS_PER_DAY})*{K('idle')},0)"),
        "your_nodes": lambda c: f"=ROUNDUP({c}{D('your_pc')}/{c}{D('cores')},0)",
        "your_util": lambda c: (f'=IF({c}{D("capacity")}="","",'
                                f'{c}{D("your_ch")}/{c}{D("capacity")})'),
    }
    for r in spec["computed"]:
        ws.cell(row=D(r["key"]), column=1, value=r["label"]).font = BLACK
        for c in "BCDE":
            cell = ws.cell(row=D(r["key"]), column="BCDE".index(c) + 2,
                           value=formulas[r["key"]](c))
            cell.number_format = FORMATS[r["format"]]
    ws.cell(row=row + 1, column=1, value=check_note(spec["note"])).font = ITALIC


def render_pricing(wb, spec, P, plan, n_loy, fy0, milestones,
                   K, C, H, R, A, Q, AU, QS, last_col):
    ws = wb.create_sheet(spec["tab"])
    ws.sheet_properties.tabColor = spec["tab_color"]
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 11
    for i in range(n_loy):
        ws.column_dimensions[col(i)].width = 13
    total_col = 3 + n_loy
    ws["A1"] = spec["title"]
    ws["A1"].font = TITLE
    ws.cell(row=2, column=1, value="Metric").font = BOLD
    ws.cell(row=2, column=2, value="Unit").font = BOLD
    for i in range(n_loy):
        ws.cell(row=2, column=3 + i, value=f"FY{fy0 + i}").font = BOLD
        milestone = milestones.get(i + 1, "")
        ws.cell(row=3, column=3 + i,
                value=(f"{milestone} (actual)" if i == 0 else milestone)).font = ITALIC
    ws.cell(row=2, column=total_col, value=spec["total_header"]).font = BOLD
    for c in range(1, total_col + 1):
        ws.cell(row=2, column=c).fill = HDR_FILL

    history = P["purchase_history"]
    hist_fy0 = history["first_fiscal_year"]
    hist_last = hist_fy0 + len(history["batch_nodes"]) - 1
    life = C("life_c")

    def infl(c):
        return f"(1+{C('fac')})^({idx(c)})"

    def cohort(hist_key, own_row):
        """Historic cohort due this FY, plus re-buy of this tab's own purchases."""
        return (lambda c:
                f"=IF(AND(COLUMN()+{fy0}-3-{life}>={hist_fy0},"
                f"COLUMN()+{fy0}-3-{life}<={hist_last}),"
                f"INDEX({H(hist_key)},1,COLUMN()+{fy0}-3-{life}-{hist_fy0 - 1}),0)"
                f"+IF(COLUMN()-2>{life},"
                f"INDEX($C${own_row}:${last_col}${own_row},1,COLUMN()-2-{life}),0)")

    fleet = P["current_fleet"]
    base_nodes = (fleet["batch"]["milano_nodes"] + fleet["batch"]["torino_nodes"]
                  + fleet["k8s"]["nodes"] + P["qserv"]["current_nodes"]
                  + fleet["interactive"]["nodes"] + fleet["cassandra"]["nodes"])

    formulas = {
        "d_years": lambda c: f"={AU}!{c}{A('survey_years')}",
        "d_ch": lambda c: f"={AU}!{c}{A('core_hours')}",
        "d_cores": lambda c: f"={AU}!{c}{A('peak_cores')}",
        "b_cum": lambda c: (f"=MAX(0,ROUNDUP(({c}{R('d_cores')}-{K('milano_eq')})"
                            f"/({K('cores_node')}*{K('speedup')}),0))"),
        "b_growth": lambda c: f"={c}{R('b_cum')}-{prev(c)}{R('b_cum')}",
        "b_refresh": cohort("hist_batch", R("b_nodes")),
        "b_nodes": lambda c: f"={c}{R('b_growth')}+{c}{R('b_refresh')}",
        "b_cost": lambda c: f"={c}{R('b_nodes')}*{C('torino')}*{infl(c)}",
        "k_need": lambda c: f"={AU}!{c}{A('onfloor_tb')}",
        "k_todate": lambda c: f"={prev(c)}{R('k_todate')}+{c}{R('k_buy')}",
        "k_buy": lambda c: f"=MAX(0,{c}{R('k_need')}-{prev(c)}{R('k_todate')})",
        "k_rebuy": lambda c: (f"=IF(COLUMN()-2>{C('life_s')},"
                              f"INDEX($C${R('k_buy')}:${last_col}${R('k_buy')},"
                              f"1,COLUMN()-2-{C('life_s')}),0)"),
        "k_cost": lambda c: (f"=({c}{R('k_buy')}+{c}{R('k_rebuy')})"
                             f"*{C('ceph')}*{infl(c)}"),
        "t_add": lambda c: f"={AU}!{c}{A('tape_buy')}",
        "t_cost": lambda c: f"={c}{R('t_add')}*{C('tape')}*{infl(c)}",
        "q_k8s": lambda c: (f"=ROUND({K('k8s_nodes')}*{C('k8s_g')},0)"
                            f"+IF(AND(COLUMN()+{fy0}-3-{life}>={hist_fy0},"
                            f"COLUMN()+{fy0}-3-{life}<={hist_last}),"
                            f"INDEX({H('hist_k8s')},1,"
                            f"COLUMN()+{fy0}-3-{life}-{hist_fy0 - 1}),0)"
                            f"+IF(COLUMN()-2>{life},"
                            f"INDEX($C${R('q_k8s')}:${last_col}${R('q_k8s')},"
                            f"1,COLUMN()-2-{life}),0)"),
        "q_k8s_cost": lambda c: f"={c}{R('q_k8s')}*{C('k8snode')}*{infl(c)}",
        "q_need": lambda c: f"={QS}!{c}{Q('nodes_req')}",
        "q_max": lambda c: f"=MAX({prev(c)}{R('q_max')},{c}{R('q_need')})",
        "q_growth": lambda c: f"={c}{R('q_max')}-{prev(c)}{R('q_max')}",
        "q_refresh": cohort("hist_qserv", R("q_nodes")),
        "q_nodes": lambda c: f"={c}{R('q_growth')}+{c}{R('q_refresh')}",
        "q_cost": lambda c: f"={c}{R('q_nodes')}*{C('k8snode')}*{infl(c)}",
        "hw": lambda c: (f"={c}{R('b_cost')}+{c}{R('k_cost')}+{c}{R('t_cost')}"
                         f"+{c}{R('q_k8s_cost')}+{c}{R('q_cost')}"),
        "nodes_svc": lambda c: (f"={base_nodes}"
                                f"+SUM($C${R('b_growth')}:{c}{R('b_growth')})"
                                f"+SUM($C${R('q_growth')}:{c}{R('q_growth')})"
                                f"+SUM($C${R('q_k8s')}:{c}{R('q_k8s')})"),
        "hosting": lambda c: f"={c}{R('nodes_svc')}*{C('host')}",
        "overhead": lambda c: f"={c}{R('hw')}*{C('ovh')}+{C('ovhf')}",
        "total": lambda c: f"={c}{R('hw')}+{c}{R('hosting')}+{c}{R('overhead')}",
        "total_m": lambda c: f"={c}{R('total')}",
        "cumulative": lambda c: f"={prev(c)}{R('cumulative')}+{c}{R('total_m')}",
    }
    firsts = {
        "b_growth": f"={C('act_batch')}",
        "k_todate": f"={K('installed_pb')}*{TB_PER_PB}+C{R('k_buy')}",
        "k_buy": f"={C('act_disk_tb')}",
        "k_rebuy": 0,
        "k_cost": f"={C('act_disk_c')}",
        "t_add": f"={C('act_arc_tb')}",
        "t_cost": f"={C('act_arc_c')}",
        "q_k8s": f"={C('act_k8s_n')}",
        "q_k8s_cost": f"={C('act_k8s_c')}",
        "q_max": f"=MAX({K('qs_nodes')},C{R('q_need')})",
        "q_growth": f"=MAX(0,C{R('q_need')}-{K('qs_nodes')})",
        "cumulative": f"=C{R('total_m')}",
    }
    render_projection(ws, spec, P, plan, n_loy, formulas, firsts,
                      total_col=total_col)


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Build the data-facility sizing and pricing workbook.")
    parser.add_argument("-p", "--params", default=DEFAULT_PARAMS,
                        help=f"parameter file (default: {DEFAULT_PARAMS})")
    parser.add_argument("-o", "--output", default=None,
                        help="output workbook (default: workbook.output in the YAML)")
    args = parser.parse_args(argv)

    params_path = Path(args.params)
    if not params_path.exists():
        parser.error(f"parameter file not found: {params_path}")
    with params_path.open() as handle:
        P = yaml.safe_load(handle)

    output = build_workbook(P, args.output)
    print(f"Wrote {output}")
    print("Formulas are written without cached values. Open in Excel or "
          "LibreOffice once to populate them.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
